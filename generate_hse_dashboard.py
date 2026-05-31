#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
 ASANKO-STYLE GOLD MINE  |  ENTERPRISE HSE MANAGEMENT DASHBOARD GENERATOR
================================================================================
A single, self-contained Python script that builds a premium, multi-sheet,
*live* Excel HSE (Health, Safety & Environment) management dashboard using only
openpyxl + pandas + numpy (no internet / no exotic dependencies).

WHY THIS IS "LIVE" (NOT A STATIC VALUE DUMP)
--------------------------------------------
There is ONE auditable data spine -- the `Incident_Register` and `Activity_Log`
sheets (plus a `Settings` sheet of rates/targets/thresholds). EVERY KPI, status
light, chart and roll-up on every other sheet is a native Excel FORMULA
(SUMPRODUCT / SUMIFS / COUNTIFS / AVERAGEIFS / TODAY / EDATE ...) that points
back at that spine. Excel Tables auto-expand and structured references
auto-grow, so when a user appends a new row to a data sheet, every number,
traffic-light and chart on the workbook recalculates automatically -- with no
redesign and no re-running this script.

Interactive Year / Month / Department / Location dropdowns on the Dashboard
drive the headline KPI formulas (with full "All" handling), and the charts read
from a filter-aware calculation engine, so the whole workbook responds to the
selected filter.

The seeded sample data (fixed numpy seed) only exists so the file opens already
populated -- delete it and type your own and the workbook keeps working.

Author: generated for an Asanko-style gold mine (Ghana) HSE programme.
================================================================================
"""

from __future__ import annotations

import calendar
import datetime as dt
from dataclasses import dataclass, field

import numpy as np
import pandas as pd

import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

# =============================================================================
# 1. CONFIGURATION, CORPORATE THEME, ENUMERATIONS
# =============================================================================

OUTPUT_FILE = "Asanko_HSE_Management_Dashboard.xlsx"
COMPANY_NAME = "ASANKO GOLD MINE"
COMPANY_SUB = "Health, Safety & Environment Management System"
RNG_SEED = 42
MONTHS_OF_HISTORY = 24          # trailing months ending at REPORT_ANCHOR
REPORT_ANCHOR = dt.date(2026, 5, 31)   # "today" the seed data is built around

# ----- Corporate colour palette (hex, no leading '#') -----------------------
NAVY = "0F2A43"      # deep header navy
NAVY2 = "1B3A5B"     # secondary navy
GOLD = "C9A227"      # mine gold accent
GOLD_L = "E6C65A"    # light gold
SLATE = "33475B"     # body text
GREY = "6B7280"      # muted label grey
LIGHT = "F2F4F7"     # light panel background
PANEL = "FFFFFF"     # card background
BORDER_GREY = "D6DCE5"
BAND = "EAEFF5"      # alt band

# Traffic-light fills / fonts
GREEN_BG, GREEN_TX = "C6EFCE", "1A7A3D"
AMBER_BG, AMBER_TX = "FFEB9C", "8A6D00"
RED_BG, RED_TX = "FFC7CE", "9C0006"
BLUE_BG, BLUE_TX = "DDEBF7", "1F4E79"

# Risk matrix band colours
RISK_LOW = "C6EFCE"
RISK_MED = "FFEB9C"
RISK_HIGH = "FFC000"
RISK_CRIT = "FF6B6B"

# Chart series colours
C_SERIES = ["1F4E79", GOLD, "2E8B57", "C0504D", "8064A2", "4BACC6", "F79646", "9BBB59"]

# Number formats
FMT_INT = "#,##0"
FMT_DEC2 = "0.00"
FMT_PCT0 = "0%"
FMT_PCT1 = "0.0%"
FMT_DATE = "dd-mmm-yyyy"
FMT_MMM = "mmm-yy"

# ----- HSE domain enumerations ----------------------------------------------
INCIDENT_TYPES = [
    "Medical Treatment", "First Aid", "Restricted Work", "Lost Time Injury",
    "Property Damage", "Environmental", "Other",
]
RECORDABLE_TYPES = ["Lost Time Injury", "Restricted Work", "Medical Treatment"]
INCIDENT_CLASSES = [
    "Personal Injury", "Occupational Illness", "Fire / Explosion",
    "Chemical / Cyanide Spill", "Vehicle / Mobile Equipment", "Slip / Trip / Fall",
    "Fall from Height", "Caught Between / Struck By", "Electrical",
    "Environmental Release", "Security", "Other",
]
INCIDENT_STATUS = ["Open", "Under Investigation", "Action Pending", "Closed"]
ACTION_STATUS = ["Open", "In Progress", "Due Soon", "Closed"]
ADEQUACY = ["Adequate", "Needs Improvement", "Inadequate"]
PRIORITY = ["Low", "Medium", "High", "Critical"]
AUDIT_TYPES = ["Internal", "External", "Regulatory"]

# Area -> (Department, dominant Company) mapping. This is the master location
# list reused everywhere so SUMPRODUCT equality matches exactly across sheets.
AREA_INFO = [
    ("Nkran Open Pit",            "Mining",         "AUMS (Mining Contractor)"),
    ("Esaase Open Pit",           "Mining",         "AUMS (Mining Contractor)"),
    ("Drill & Blast",             "Mining",         "Maxam (Blasting)"),
    ("Haul Roads",                "Mining",         "AUMS (Mining Contractor)"),
    ("Explosives Magazine",       "Mining",         "Maxam (Blasting)"),
    ("Crushing Circuit",          "Processing",     "Owner (Asanko)"),
    ("Processing Plant (CIL)",    "Processing",     "Owner (Asanko)"),
    ("Elution & Gold Room",       "Processing",     "Owner (Asanko)"),
    ("Tailings Storage Facility", "Environment",    "Owner (Asanko)"),
    ("Water Treatment",           "Environment",    "Owner (Asanko)"),
    ("Assay Laboratory",          "Laboratory",     "SGS (Laboratory)"),
    ("HME Workshop",              "Engineering",    "Owner (Asanko)"),
    ("Fuel Farm",                 "Engineering",    "Owner (Asanko)"),
    ("Power Station",             "Engineering",    "Genser (Power)"),
    ("Warehouse & Stores",        "Logistics",      "Owner (Asanko)"),
    ("Administration",            "Administration", "Owner (Asanko)"),
    ("Accommodation Camp",        "Administration", "Catering Co (Camp)"),
    ("Security Gatehouse",        "Security",       "G4S (Security)"),
]
AREAS = [a[0] for a in AREA_INFO]
DEPARTMENTS = sorted({a[1] for a in AREA_INFO})
COMPANIES = ["Owner (Asanko)", "AUMS (Mining Contractor)", "Maxam (Blasting)",
             "SGS (Laboratory)", "Genser (Power)", "Catering Co (Camp)", "G4S (Security)"]
AREA_DEPT = {a[0]: a[1] for a in AREA_INFO}
AREA_COMPANY = {a[0]: a[2] for a in AREA_INFO}

OWNERS = ["K. Mensah", "A. Owusu", "J. Boateng", "E. Asante", "P. Annan",
          "Y. Darko", "S. Addo", "M. Quaye", "F. Agyeman", "D. Tetteh"]

# Settings: (defined_name, label, value, number_format). Formulas reference the
# defined names -- never the literal numbers.
SETTINGS = [
    ("RATE_BASE",            "Frequency-rate base (mining standard)",      1_000_000, FMT_INT),
    ("RATE_BASE_OSHA",       "Frequency-rate base (OSHA 200,000 std)",       200_000, FMT_INT),
    ("TARGET_TRIFR",         "Target TRIFR (per 1,000,000 hrs)",                 3.0, FMT_DEC2),
    ("TARGET_LTIFR",         "Target LTIFR (per 1,000,000 hrs)",                 0.8, FMT_DEC2),
    ("TARGET_INSPECTION",    "Target inspection performance score",             0.95, FMT_PCT0),
    ("TARGET_TRAINING",      "Target training completion rate",                 0.90, FMT_PCT0),
    ("TARGET_ENV",           "Target environmental compliance",                 0.98, FMT_PCT0),
    ("TARGET_NEARMISS",      "Target near-miss reports (per month)",              40, FMT_INT),
    ("NEARMISS_RATIO_TGT",   "Target near-miss : recordable ratio",               10, FMT_INT),
    ("THRESH_LTI_GOOD",      "Days-since-LTI: green threshold",                   30, FMT_INT),
    ("THRESH_LTI_WARN",      "Days-since-LTI: amber threshold",                    7, FMT_INT),
    ("PM10_LIMIT",           "PM10 dust limit (ug/m3, EPA Ghana)",                70, FMT_INT),
    ("PH_MIN",               "Discharge pH lower limit",                         6.0, FMT_DEC2),
    ("PH_MAX",               "Discharge pH upper limit",                         9.0, FMT_DEC2),
    ("WADCN_LIMIT",          "WAD cyanide limit (mg/L, ICMC)",                    50, FMT_INT),
]

# Fixed cell anchors so defined names can be created up-front (decoupled from
# the order sheets are built in).
ADDR_FILTER_YEAR = "Dashboard!$C$6"
ADDR_FILTER_MONTH = "Dashboard!$F$6"
ADDR_FILTER_DEPT = "Dashboard!$I$6"
ADDR_FILTER_LOC = "Dashboard!$L$6"
ADDR_ENV_PCT = "Environmental!$C$4"

# =============================================================================
# 2. LOW-LEVEL STYLING HELPERS
# =============================================================================

THIN = Side(style="thin", color=BORDER_GREY)
MED = Side(style="medium", color=NAVY)
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_BOX = Border(left=MED, right=MED, top=MED, bottom=MED)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def fill(hex_color: str) -> PatternFill:
    """Solid fill shortcut."""
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def coord(row: int, col: int) -> str:
    """(row, col) -> 'A1' style address."""
    return f"{get_column_letter(col)}{row}"


def style_range(ws, cell_range, *, fillc=None, font=None, align=None, border=None,
                number_format=None):
    """Apply formatting to every cell in a range (needed for merged regions)."""
    for row in ws[cell_range]:
        for c in row:
            if fillc is not None:
                c.fill = fillc
            if font is not None:
                c.font = font
            if align is not None:
                c.alignment = align
            if border is not None:
                c.border = border
            if number_format is not None:
                c.number_format = number_format


def set_widths(ws, widths: dict):
    """widths: {'A': 12, 'B': 20, ...}"""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def add_defined_name(wb, name: str, ref: str):
    """Create / overwrite a workbook-level defined name (openpyxl 3.1 dict API)."""
    wb.defined_names[name] = DefinedName(name, attr_text=ref)


def title_band(ws, title, subtitle, last_col_letter, *, with_date=True):
    """Dark navy title band with gold accent rule used at the top of every sheet."""
    ws.merge_cells(f"A1:{last_col_letter}2")
    style_range(ws, f"A1:{last_col_letter}2", fillc=fill(NAVY), align=Alignment(
        horizontal="left", vertical="center", indent=1))
    t = ws["A1"]
    t.value = f"{COMPANY_NAME}   |   {title}"
    t.font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
    # gold accent rule
    ws.merge_cells(f"A3:{last_col_letter}3")
    style_range(ws, f"A3:{last_col_letter}3", fillc=fill(GOLD))
    s = ws["A3"]
    s.value = subtitle
    s.font = Font(name="Calibri", size=10, bold=True, color=NAVY)
    s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    if with_date:
        dcell = coord(3, ws.max_column if ws.max_column > 4 else 8)
    ws.row_dimensions[1].height = 16
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 18


def section_header(ws, cell_range, text):
    """A slim navy section sub-header bar."""
    ws.merge_cells(cell_range)
    first = cell_range.split(":")[0]
    style_range(ws, cell_range, fillc=fill(NAVY2),
                align=Alignment(horizontal="left", vertical="center", indent=1))
    ws[first].value = text
    ws[first].font = Font(bold=True, color="FFFFFF", size=11)


# =============================================================================
# 3. FILTER-AWARE PREDICATE BUILDERS  (the heart of "everything is live")
# =============================================================================
# Each filter term resolves to 1 when the filter cell is "All" / blank OR when
# the row's value matches the selected filter -- giving robust "All" handling
# inside SUMPRODUCT.

def _term(col_ref: str, filter_cell: str) -> str:
    return f'(({filter_cell}="All")+({filter_cell}="")+({col_ref}={filter_cell}))'


def pred(table: str, *, year=True, month=True, dept=True, loc=True) -> str:
    """Build a SUMPRODUCT filter predicate for a given Excel table."""
    parts = []
    if year:
        parts.append(_term(f"{table}[Year]", "F_Year"))
    if month:
        parts.append(_term(f"{table}[MonthName]", "F_Month"))
    if dept:
        parts.append(_term(f"{table}[Department]", "F_Dept"))
    if loc:
        parts.append(_term(f"{table}[Area]", "F_Loc"))
    return "*".join(parts)


# =============================================================================
# 4. GENERIC DATAFRAME -> EXCEL TABLE WRITER
# =============================================================================

@dataclass
class Layout:
    """Where a written table lives, so callers can build chart refs / CF ranges."""
    ws: object
    top_row: int                 # header row
    left_col: int
    n_rows: int                  # data rows
    n_cols: int
    col_index: dict = field(default_factory=dict)   # header -> absolute column #

    @property
    def first_data_row(self):
        return self.top_row + 1

    @property
    def last_data_row(self):
        return self.top_row + self.n_rows

    def col(self, header):
        return self.col_index[header]

    def col_letter(self, header):
        return get_column_letter(self.col_index[header])

    def data_range(self, header, buffer=0):
        """e.g. 'K4:K27' for the data body of a column (optionally + buffer)."""
        c = self.col_letter(header)
        return f"{c}{self.first_data_row}:{c}{self.last_data_row + buffer}"


def _coerce(v):
    """Make numpy / pandas scalars safe for openpyxl."""
    if v is None:
        return None
    if isinstance(v, float) and pd.isna(v):
        return None
    if isinstance(v, (np.integer,)):
        return int(v)
    if isinstance(v, (np.floating,)):
        return float(v)
    if isinstance(v, (np.bool_,)):
        return bool(v)
    if isinstance(v, pd.Timestamp):
        return v.to_pydatetime().date()
    return v


def write_table(ws, df: pd.DataFrame, top_row: int, left_col: int, table_name: str,
                *, number_formats: dict | None = None, style="TableStyleMedium2",
                header_fill=NAVY, header_font="FFFFFF") -> Layout:
    """Write a DataFrame as a native Excel Table (auto-expanding, with filters)."""
    number_formats = number_formats or {}
    headers = list(df.columns)
    col_index = {h: left_col + j for j, h in enumerate(headers)}

    # header row
    for j, h in enumerate(headers):
        c = ws.cell(top_row, left_col + j, h)
        c.fill = fill(header_fill)
        c.font = Font(bold=True, color=header_font, size=10)
        c.alignment = CENTER
        c.border = BORDER_THIN
    ws.row_dimensions[top_row].height = 28

    # body
    for i, (_, r) in enumerate(df.iterrows()):
        excel_row = top_row + 1 + i
        for j, h in enumerate(headers):
            c = ws.cell(excel_row, left_col + j, _coerce(r[h]))
            c.border = BORDER_THIN
            c.alignment = LEFT if isinstance(r[h], str) else CENTER
            if h in number_formats:
                c.number_format = number_formats[h]

    # define the Excel Table (header + data extent)
    ref = f"{coord(top_row, left_col)}:{coord(top_row + len(df), left_col + len(headers) - 1)}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name=style, showRowStripes=True, showColumnStripes=False,
        showFirstColumn=False, showLastColumn=False)
    ws.add_table(table)

    return Layout(ws, top_row, left_col, len(df), len(headers), col_index)


# =============================================================================
# 5. KPI CARD + TRAFFIC-LIGHT HELPERS (Dashboard)
# =============================================================================

def kpi_card(ws, col, row, *, title, formula, fmt=FMT_INT, value_size=24,
             accent=GOLD, suffix_formula=None):
    """
    Draw a modern KPI 'card' (3 cols x 4 rows):
        row   : title label (small, grey)
        +1,+2 : big value (merged) -- a live formula
        +3    : status / sub-text line
    Returns (value_cell_addr, status_cell_addr) so callers can attach CF.
    """
    c0, c1 = col, col + 2
    L = get_column_letter
    title_rng = f"{L(c0)}{row}:{L(c1)}{row}"
    val_rng = f"{L(c0)}{row + 1}:{L(c1)}{row + 2}"
    stat_rng = f"{L(c0)}{row + 3}:{L(c1)}{row + 3}"
    whole = f"{L(c0)}{row}:{L(c1)}{row + 3}"

    # card background + border + left accent bar
    style_range(ws, whole, fillc=fill(PANEL), border=BORDER_THIN)
    style_range(ws, f"{L(c0)}{row}:{L(c0)}{row + 3}", fillc=fill(accent))

    ws.merge_cells(title_rng)
    ws.merge_cells(val_rng)
    ws.merge_cells(stat_rng)

    tc = ws[f"{L(c0)}{row}"]
    tc.value = title
    tc.font = Font(bold=True, size=9, color=GREY)
    tc.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    vc = ws[f"{L(c0)}{row + 1}"]
    vc.value = formula
    vc.number_format = fmt
    vc.font = Font(bold=True, size=value_size, color=NAVY)
    vc.alignment = Alignment(horizontal="center", vertical="center")

    sc = ws[f"{L(c0)}{row + 3}"]
    if suffix_formula is not None:
        sc.value = suffix_formula
    sc.font = Font(bold=True, size=9, color="FFFFFF")
    sc.alignment = CENTER

    return f"{L(c0)}{row + 1}", f"{L(c0)}{row + 3}"


def traffic_light(ws, value_addr, status_addr, mode, target_ref=None):
    """Attach conditional-format traffic lights to a KPI card's status cell."""
    g = (fill(GREEN_BG), Font(bold=True, color=GREEN_TX, size=9))
    a = (fill(AMBER_BG), Font(bold=True, color=AMBER_TX, size=9))
    r = (fill(RED_BG), Font(bold=True, color=RED_TX, size=9))
    cf = ws.conditional_formatting

    def rule(formula, style):
        return FormulaRule(formula=[formula], fill=style[0], font=style[1], stopIfTrue=True)

    if mode == "lower_better":            # TRIFR, LTIFR
        cf.add(status_addr, rule(f"{value_addr}<={target_ref}", g))
        cf.add(status_addr, rule(f"{value_addr}<={target_ref}*1.15", a))
        cf.add(status_addr, rule(f"{value_addr}>{target_ref}*1.15", r))
    elif mode == "higher_better":         # inspection, training, env, near-miss
        cf.add(status_addr, rule(f"{value_addr}>={target_ref}", g))
        cf.add(status_addr, rule(f"{value_addr}>={target_ref}*0.9", a))
        cf.add(status_addr, rule(f"{value_addr}<{target_ref}*0.9", r))
    elif mode == "zero_best":             # overdue actions
        cf.add(status_addr, rule(f"{value_addr}=0", g))
        cf.add(status_addr, rule(f"{value_addr}>0", r))
    elif mode == "days_lti":              # days since last LTI
        cf.add(status_addr, rule(f"{value_addr}>=THRESH_LTI_GOOD", g))
        cf.add(status_addr, rule(f"{value_addr}>=THRESH_LTI_WARN", a))
        cf.add(status_addr, rule(f"{value_addr}<THRESH_LTI_WARN", r))


def status_cf(ws, cell_range, mapping):
    """Colour a Status column by exact text value."""
    fonts = {GREEN_BG: GREEN_TX, AMBER_BG: AMBER_TX, RED_BG: RED_TX, BLUE_BG: BLUE_TX}
    for value, bg in mapping.items():
        ws.conditional_formatting.add(cell_range, CellIsRule(
            operator="equal", formula=[f'"{value}"'], fill=fill(bg),
            font=Font(bold=True, color=fonts.get(bg, SLATE))))


STATUS_MAPS = {
    "incident": {"Closed": GREEN_BG, "Open": RED_BG,
                 "Under Investigation": AMBER_BG, "Action Pending": AMBER_BG},
    "action": {"Closed": GREEN_BG, "Open": RED_BG,
               "In Progress": BLUE_BG, "Due Soon": AMBER_BG},
    "compliance": {"Compliant": GREEN_BG, "Due Soon": AMBER_BG, "Overdue": RED_BG},
    "permit": {"Active": GREEN_BG, "Expiring": AMBER_BG, "Expired": RED_BG},
    "drill": {"Current": GREEN_BG, "Due Soon": AMBER_BG, "Overdue": RED_BG},
    "equip": {"Compliant": GREEN_BG, "Due Soon": AMBER_BG, "Overdue": RED_BG},
    "ppe": {"Compliant": GREEN_BG, "Monitor": AMBER_BG, "Action": RED_BG},
    "audit": {"Closed": GREEN_BG, "Open": AMBER_BG},
    "risk": {"Low": GREEN_BG, "Medium": AMBER_BG, "High": RISK_HIGH, "Critical": RISK_CRIT},
}


# =============================================================================
# 6. CHART HELPERS
# =============================================================================

def _style_chart(ch, title):
    ch.title = title
    ch.style = 2
    ch.height = 7.4
    ch.width = 15.5
    if ch.title and hasattr(ch.title, "tx"):
        pass
    return ch


def line_chart(ws_src, title, cats_ref, data_refs, anchor, ws_dst, *,
               y_title="", colours=None):
    ch = LineChart()
    _style_chart(ch, title)
    ch.y_axis.title = y_title
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    for i, dref in enumerate(data_refs):
        ch.add_data(dref, titles_from_data=True)
    ch.set_categories(cats_ref)
    for i, s in enumerate(ch.series):
        s.smooth = False
        s.graphicalProperties.line.width = 28000
        col = (colours or C_SERIES)[i % len(C_SERIES)]
        s.graphicalProperties.line.solidFill = col
    ws_dst.add_chart(ch, anchor)
    return ch


def bar_chart(ws_src, title, cats_ref, data_ref, anchor, ws_dst, *,
              bar_dir="col", colour=NAVY2, y_title="", data_labels=False):
    ch = BarChart()
    _style_chart(ch, title)
    ch.type = bar_dir
    ch.y_axis.title = y_title
    ch.legend = None
    ch.add_data(data_ref, titles_from_data=True)
    ch.set_categories(cats_ref)
    ch.series[0].graphicalProperties.solidFill = colour
    ch.gapWidth = 60
    if data_labels:
        ch.dataLabels = DataLabelList()
        ch.dataLabels.showVal = True
    ws_dst.add_chart(ch, anchor)
    return ch


def doughnut_chart(ws_src, title, cats_ref, data_ref, anchor, ws_dst, colours=None):
    ch = DoughnutChart()
    ch.title = title
    ch.height = 7.4
    ch.width = 8.6
    ch.holeSize = 55
    ch.add_data(data_ref, titles_from_data=True)
    ch.set_categories(cats_ref)
    ch.dataLabels = DataLabelList()
    ch.dataLabels.showPercent = True
    ws_dst.add_chart(ch, anchor)
    return ch


def combo_chart(ws_src, title, cats_ref, bar_data, line_data, anchor, ws_dst, *,
                bar_title="", line_title=""):
    """Clustered bars (primary axis) + line (secondary axis) -- leading vs lagging."""
    bar = BarChart()
    _style_chart(bar, title)
    bar.type = "col"
    bar.grouping = "clustered"
    bar.y_axis.title = bar_title
    bar.add_data(bar_data, titles_from_data=True)
    bar.set_categories(cats_ref)
    for i, s in enumerate(bar.series):
        s.graphicalProperties.solidFill = C_SERIES[i % len(C_SERIES)]
    bar.gapWidth = 80

    line = LineChart()
    line.add_data(line_data, titles_from_data=True)
    line.set_categories(cats_ref)
    line.y_axis.axId = 200
    line.y_axis.title = line_title
    line.y_axis.crosses = "max"
    for s in line.series:
        s.graphicalProperties.line.solidFill = "C0504D"
        s.graphicalProperties.line.width = 30000
        s.smooth = False

    bar += line
    ws_dst.add_chart(bar, anchor)
    return bar


# =============================================================================
# 7. SAMPLE DATA GENERATION  (fixed seed -> reproducible, fully editable)
# =============================================================================

def month_timeline(anchor: dt.date, n: int):
    """Return list of dicts for the trailing `n` months ending in `anchor`."""
    months = []
    y, m = anchor.year, anchor.month
    for _ in range(n):
        months.append({"year": y, "month": m, "abbr": calendar.month_abbr[m],
                       "first": dt.date(y, m, 1)})
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    return list(reversed(months))


def generate_sample_data():
    """Build every seed DataFrame. Returns a dict of DataFrames + the timeline."""
    rng = np.random.default_rng(RNG_SEED)
    months = month_timeline(REPORT_ANCHOR, MONTHS_OF_HISTORY)
    today = REPORT_ANCHOR

    # ---- Incident_Register --------------------------------------------------
    # Recordables (MTC/RWC/LTI) are deliberately a small slice of all reported
    # events so frequency rates land in a realistic band near the targets.
    # Order matches INCIDENT_TYPES: MTC, First Aid, RWC, LTI, Prop.Dmg, Env, Other
    type_w = np.array([0.045, 0.42, 0.025, 0.02, 0.16, 0.15, 0.18])
    type_w = type_w / type_w.sum()
    inc_rows = []
    inc_id = 1
    for k, mo in enumerate(months):
        # improving trend: mean incidents/month declines gently over time
        lam = 9.0 - 4.0 * (k / max(1, len(months) - 1))
        n_inc = int(rng.poisson(max(2.0, lam)))
        ndays = calendar.monthrange(mo["year"], mo["month"])[1]
        for _ in range(n_inc):
            area = rng.choice(AREAS)
            itype = rng.choice(INCIDENT_TYPES, p=type_w)
            # severity correlates with type
            if itype == "Lost Time Injury":
                sev = int(rng.choice([3, 4, 5], p=[0.5, 0.35, 0.15]))
            elif itype in ("Restricted Work", "Property Damage", "Environmental"):
                sev = int(rng.choice([2, 3, 4], p=[0.4, 0.45, 0.15]))
            else:
                sev = int(rng.choice([1, 2, 3], p=[0.55, 0.35, 0.10]))
            day = int(rng.integers(1, ndays + 1))
            date = dt.date(mo["year"], mo["month"], day)
            reported = date + dt.timedelta(days=int(rng.integers(0, 3)))
            age = (today - date).days
            if age > 150:
                status = rng.choice(INCIDENT_STATUS, p=[0.02, 0.03, 0.05, 0.90])
            elif age > 60:
                status = rng.choice(INCIDENT_STATUS, p=[0.08, 0.12, 0.20, 0.60])
            else:
                status = rng.choice(INCIDENT_STATUS, p=[0.30, 0.30, 0.20, 0.20])
            needs_car = itype in RECORDABLE_TYPES or itype in ("Environmental", "Property Damage")
            car_due = reported + dt.timedelta(days=int(rng.integers(14, 61))) if needs_car else None
            inc_rows.append({
                "ID": f"INC-{inc_id:04d}", "Date": date,
                "Year": f"=YEAR([@Date])", "MonthName": f'=TEXT([@Date],"mmm")',
                "Month": f"=MONTH([@Date])",
                "Area": area, "Department": AREA_DEPT[area], "Company": AREA_COMPANY[area],
                "Type": itype, "Class": rng.choice(INCIDENT_CLASSES), "Severity": sev,
                "Status": status, "Reported": reported, "CAR_Due": car_due,
                "Owner": rng.choice(OWNERS),
                "Recordable": ('=IF(OR([@Type]="Lost Time Injury",[@Type]="Restricted Work",'
                               '[@Type]="Medical Treatment"),1,0)'),
                "LTI": '=IF([@Type]="Lost Time Injury",1,0)',
                "LTI_Date": '=IF([@Type]="Lost Time Injury",[@Date],"")',
                "CAR_Overdue": ('=IF(AND([@Status]<>"Closed",[@CAR_Due]<>"",'
                                '[@CAR_Due]<TODAY()),1,0)'),
                "Days_Overdue": ('=IF(AND([@Status]<>"Closed",[@CAR_Due]<>"",'
                                 '[@CAR_Due]<TODAY()),TODAY()-[@CAR_Due],0)'),
            })
            inc_id += 1
    incidents = pd.DataFrame(inc_rows)

    # ---- Activity_Log (per area, per month) --------------------------------
    base_head = {  # rough relative headcount/exposure per area
        "Nkran Open Pit": 120, "Esaase Open Pit": 110, "Drill & Blast": 40,
        "Haul Roads": 60, "Explosives Magazine": 15, "Crushing Circuit": 35,
        "Processing Plant (CIL)": 90, "Elution & Gold Room": 20,
        "Tailings Storage Facility": 18, "Water Treatment": 12, "Assay Laboratory": 25,
        "HME Workshop": 55, "Fuel Farm": 10, "Power Station": 22,
        "Warehouse & Stores": 24, "Administration": 60, "Accommodation Camp": 30,
        "Security Gatehouse": 45,
    }
    act_rows = []
    for mo in months:
        for area in AREAS:
            hc = base_head[area]
            manhours = int(hc * rng.integers(180, 230))           # ~ monthly hours
            nm = int(rng.poisson(max(1, hc / 22)))
            hz = int(rng.poisson(max(1, hc / 30)))
            obs_r = int(rng.poisson(max(2, hc / 12)))
            obs_c = int(round(obs_r * rng.uniform(0.75, 0.98)))
            insp_p = int(max(2, hc / 18))
            insp_d = int(round(insp_p * rng.uniform(0.85, 1.0)))
            insp_score = round(float(rng.uniform(0.88, 0.995)), 3)
            audits = int(rng.integers(0, 2))
            toolbox = int(rng.poisson(max(2, hc / 10)))
            tr_assigned = int(max(3, hc / 6))
            tr_completed = int(round(tr_assigned * rng.uniform(0.9, 1.0)))
            ppe = round(float(rng.uniform(0.88, 1.0)), 3)
            act_rows.append({
                "Period": mo["first"], "Year": "=YEAR([@Period])",
                "MonthName": '=TEXT([@Period],"mmm")', "Month": "=MONTH([@Period])",
                "Area": area, "Department": AREA_DEPT[area], "Company": AREA_COMPANY[area],
                "ManHours": manhours, "NearMisses": nm, "Hazards": hz,
                "ObsRaised": obs_r, "ObsClosed": obs_c,
                "InspPlanned": insp_p, "InspDone": insp_d, "InspScore": insp_score,
                "Audits": audits, "Toolbox": toolbox,
                "TrainAssigned": tr_assigned, "TrainCompleted": tr_completed, "PPE": ppe,
                "InspCompletion": "=IFERROR([@InspDone]/[@InspPlanned],0)",
                "ObsCloseRate": "=IFERROR([@ObsClosed]/[@ObsRaised],0)",
            })
    activity = pd.DataFrame(act_rows)

    # ---- Corrective_Actions -------------------------------------------------
    act2_rows = []
    aid = 1
    # actions sourced from incidents needing a CAR
    car_incidents = incidents[incidents["CAR_Due"].notna()].reset_index(drop=True)
    for _, r in car_incidents.iterrows():
        raised = r["Reported"]
        due = r["CAR_Due"]
        closed = r["Status"] == "Closed"
        if closed:
            status = "Closed"
        else:
            status = rng.choice(["Open", "In Progress", "Due Soon"], p=[0.45, 0.35, 0.20])
        act2_rows.append({
            "Action_ID": f"CAR-{aid:04d}", "Source_Incident": r["ID"],
            "Description": f"Corrective action for {r['Type']} at {r['Area']}",
            "Raised": raised, "Due": due, "Owner": r["Owner"],
            "Department": r["Department"], "Area": r["Area"],
            "Priority": rng.choice(PRIORITY, p=[0.2, 0.4, 0.3, 0.1]), "Status": status,
            "Overdue": '=IF(AND([@Status]<>"Closed",[@Due]<TODAY()),1,0)',
            "Days_Overdue": '=IF(AND([@Status]<>"Closed",[@Due]<TODAY()),TODAY()-[@Due],0)',
        })
        aid += 1
    # extra proactive actions from inspections / audits
    for _ in range(28):
        area = rng.choice(AREAS)
        raised = today - dt.timedelta(days=int(rng.integers(5, 200)))
        due = raised + dt.timedelta(days=int(rng.integers(20, 75)))
        status = rng.choice(ACTION_STATUS, p=[0.20, 0.25, 0.15, 0.40])
        act2_rows.append({
            "Action_ID": f"CAR-{aid:04d}", "Source_Incident": "Inspection/Audit",
            "Description": rng.choice([
                "Install machine guard", "Repair edge protection", "Replace fire extinguisher",
                "Update SOP / JSA", "Bund repair at fuel storage", "Improve signage",
                "Spill kit replenishment", "Lighting upgrade", "Housekeeping campaign"]),
            "Raised": raised, "Due": due, "Owner": rng.choice(OWNERS),
            "Department": AREA_DEPT[area], "Area": area,
            "Priority": rng.choice(PRIORITY, p=[0.3, 0.4, 0.2, 0.1]), "Status": status,
            "Overdue": '=IF(AND([@Status]<>"Closed",[@Due]<TODAY()),1,0)',
            "Days_Overdue": '=IF(AND([@Status]<>"Closed",[@Due]<TODAY()),TODAY()-[@Due],0)',
        })
        aid += 1
    actions = pd.DataFrame(act2_rows)

    # ---- Risk register ------------------------------------------------------
    risk_defs = [
        ("Pit wall slope failure", "Nkran Open Pit", "Geotechnical", 3, 5),
        ("Haul truck / light-vehicle interaction", "Haul Roads", "Vehicle", 4, 4),
        ("Uncontrolled blast / flyrock", "Drill & Blast", "Explosives", 2, 5),
        ("Cyanide exposure (CIL)", "Processing Plant (CIL)", "Chemical", 2, 5),
        ("Tailings dam overtopping / breach", "Tailings Storage Facility", "Environmental", 2, 5),
        ("Fall from height on plant structures", "Crushing Circuit", "Personal Injury", 3, 4),
        ("Fuel / diesel spill to ground", "Fuel Farm", "Environmental", 3, 3),
        ("Confined space entry (tanks)", "Elution & Gold Room", "Personal Injury", 2, 4),
        ("Electrical arc flash", "Power Station", "Electrical", 2, 4),
        ("Heat stress / fatigue", "Open Pit operations", "Occupational Health", 4, 3),
        ("Malaria / community health", "Accommodation Camp", "Occupational Health", 4, 3),
        ("Radiation source (density gauge)", "Processing Plant (CIL)", "Radiation", 1, 4),
        ("Mobile crusher entanglement", "Crushing Circuit", "Mechanical", 2, 4),
        ("Security / illegal mining incursion", "Security Gatehouse", "Security", 3, 3),
    ]
    risk_rows = []
    for i, (desc, area, cat, L, C) in enumerate(risk_defs, start=1):
        adeq = rng.choice(ADEQUACY, p=[0.5, 0.4, 0.1])
        risk_rows.append({
            "Risk_ID": f"RR-{i:03d}", "Description": desc, "Area": area, "Category": cat,
            "Likelihood": L, "Consequence": C, "Score": "=[@Likelihood]*[@Consequence]",
            "Risk_Level": ('=IF([@Score]>=16,"Critical",IF([@Score]>=10,"High",'
                           'IF([@Score]>=5,"Medium","Low")))'),
            "Control_Adequacy": adeq, "Owner": rng.choice(OWNERS),
            "Review_Date": today + dt.timedelta(days=int(rng.integers(15, 120))),
        })
    risks = pd.DataFrame(risk_rows)

    # ---- Compliance (Ghana regulatory) -------------------------------------
    comp_defs = [
        ("Mining Lease / Operating Permit", "Minerals Commission", "Act 703", 12),
        ("Annual Mineral Right Rent", "Minerals Commission", "LI 2176", 12),
        ("Environmental Permit", "EPA Ghana", "LI 1652", 12),
        ("Annual Environmental Mgmt Report (AEMR)", "EPA Ghana", "EA Regs", 12),
        ("Effluent / Discharge Monitoring", "EPA Ghana", "GEPA Std", 3),
        ("Factory Registration & Inspection", "Factories Inspectorate", "Act 328", 12),
        ("Pressure Vessel Certification", "Factories Inspectorate", "Act 328", 12),
        ("ICMC Cyanide Code Certification", "ICMI", "Cyanide Code", 36),
        ("Cyanide Transport Audit", "ICMI", "Cyanide Code", 12),
        ("Radiation Source Licence", "Nuclear Regulatory Authority", "Act 895", 12),
        ("Water Use Permit", "Water Resources Commission", "Act 522", 12),
        ("Fire Safety Certificate", "Ghana National Fire Service", "Act 537", 12),
        ("Explosives Licence (Magazine)", "Minerals Commission", "LI 2177", 12),
    ]
    comp_rows = []
    for i, (item, reg, ref, freq) in enumerate(comp_defs, start=1):
        # spread last-completed dates so some are compliant / due soon / overdue
        last = today - dt.timedelta(days=int(rng.integers(20, int(freq * 30 + 40))))
        comp_rows.append({
            "Item": item, "Regulator": reg, "Reference": ref, "Frequency_Months": freq,
            "Last_Completed": last, "Owner": rng.choice(OWNERS),
            "Due_Date": "=EDATE([@Last_Completed],[@Frequency_Months])",
            "Status": ('=IF([@Due_Date]="","",IF(TODAY()>[@Due_Date],"Overdue",'
                       'IF([@Due_Date]-TODAY()<=30,"Due Soon","Compliant")))'),
            "Days_To_Due": "=[@Due_Date]-TODAY()",
        })
    compliance = pd.DataFrame(comp_rows)

    # ---- Environmental monitoring (monthly) --------------------------------
    env_rows = []
    for k, mo in enumerate(months):
        pm10 = round(float(rng.normal(48, 12)), 1)
        if k in (6, 17):          # a couple of dusty months exceed the limit
            pm10 = round(float(rng.uniform(74, 92)), 1)
        ph = round(float(rng.normal(7.6, 0.5)), 2)
        if k == 11:
            ph = round(float(rng.uniform(9.1, 9.6)), 2)
        cn = round(float(rng.normal(28, 9)), 1)
        if k == 20:
            cn = round(float(rng.uniform(52, 66)), 1)
        env_rows.append({
            "Period": mo["first"], "Year": "=YEAR([@Period])",
            "MonthName": '=TEXT([@Period],"mmm")',
            "Waste_t": round(float(rng.uniform(120, 260)), 1),
            "Recycling": round(float(rng.uniform(0.35, 0.62)), 3),
            "Energy_MWh": round(float(rng.uniform(5200, 6800)), 0),
            "Water_m3": round(float(rng.uniform(38000, 52000)), 0),
            "Fuel_L": round(float(rng.uniform(420000, 560000)), 0),
            "PM10": pm10, "pH": ph, "WAD_CN": cn,
            "PM10_Limit": "=PM10_LIMIT", "CN_Limit": "=WADCN_LIMIT",
            "PM10_OK": "=IF([@PM10]<=PM10_LIMIT,1,0)",
            "pH_OK": "=IF(AND([@pH]>=PH_MIN,[@pH]<=PH_MAX),1,0)",
            "CN_OK": "=IF([@WAD_CN]<=WADCN_LIMIT,1,0)",
            "Energy_Trend": ('=IF(N(OFFSET([@Energy_MWh],-1,0))=0,"",'
                             'IF([@Energy_MWh]>OFFSET([@Energy_MWh],-1,0),"UP",'
                             'IF([@Energy_MWh]<OFFSET([@Energy_MWh],-1,0),"DOWN","FLAT")))'),
            "Water_Trend": ('=IF(N(OFFSET([@Water_m3],-1,0))=0,"",'
                            'IF([@Water_m3]>OFFSET([@Water_m3],-1,0),"UP",'
                            'IF([@Water_m3]<OFFSET([@Water_m3],-1,0),"DOWN","FLAT")))'),
        })
    environmental = pd.DataFrame(env_rows)

    # ---- Contractors (man-hours & recordables live from the spine) ---------
    scope = {
        "Owner (Asanko)": "Process plant, admin, environment",
        "AUMS (Mining Contractor)": "Load & haul, drilling, pit ops",
        "Maxam (Blasting)": "Explosives supply & blasting",
        "SGS (Laboratory)": "Assay & sample preparation",
        "Genser (Power)": "Power generation & distribution",
        "Catering Co (Camp)": "Catering & camp services",
        "G4S (Security)": "Site security & access control",
    }
    contractor_rows = []
    for co in COMPANIES:
        contractor_rows.append({
            "Company": co, "Scope": scope[co],
            "ManHours": f'=SUMPRODUCT((tblActivity[Company]="{co}")*tblActivity[ManHours])',
            "Recordables": f'=SUMPRODUCT((tblIncidents[Company]="{co}")*tblIncidents[Recordable])',
            "LTIs": f'=SUMPRODUCT((tblIncidents[Company]="{co}")*tblIncidents[LTI])',
            "TRIFR": "=IF([@ManHours]=0,0,[@Recordables]/[@ManHours]*RATE_BASE)",
            "LTIFR": "=IF([@ManHours]=0,0,[@LTIs]/[@ManHours]*RATE_BASE)",
            "Target": "=TARGET_TRIFR",
        })
    contractors = pd.DataFrame(contractor_rows)

    # ---- Supporting registers ----------------------------------------------
    permits = pd.DataFrame([{
        "Permit": p, "Authority": auth, "Holder": rng.choice(OWNERS),
        "Issue_Date": today - dt.timedelta(days=int(rng.integers(120, 700))),
        "Expiry_Date": today + dt.timedelta(days=int(d)),
        "Days_To_Expiry": "=[@Expiry_Date]-TODAY()",
        "Status": ('=IF([@Expiry_Date]<TODAY(),"Expired",'
                   'IF([@Expiry_Date]-TODAY()<=60,"Expiring","Active"))'),
    } for p, auth, d in [
        ("Hot Work Permit - Plant", "HSE Dept", 45),
        ("Confined Space - Elution", "HSE Dept", 120),
        ("Working at Height - Crusher", "HSE Dept", -10),
        ("Excavation Permit - Esaase", "HSE Dept", 200),
        ("Explosives Handling", "Minerals Commission", 300),
        ("Radiation Source Permit", "Nuclear Reg. Authority", 30),
        ("Electrical Isolation - HV", "Engineering", 90),
        ("Lifting Operations Permit", "Engineering", 150),
    ]])

    audits = pd.DataFrame([{
        "Audit": a, "Type": t, "Date": today - dt.timedelta(days=int(rng.integers(10, 330))),
        "Auditor": rng.choice(OWNERS), "Score": round(float(rng.uniform(0.72, 0.97)), 3),
        "Findings": f, "Closed_Findings": int(rng.integers(0, f + 1)),
        "Status": '=IF([@Closed_Findings]>=[@Findings],"Closed","Open")',
    } for a, t, f in [
        ("ICMC Cyanide Code Audit", "External", 8),
        ("ISO 45001 Surveillance", "External", 6),
        ("EPA Environmental Audit", "Regulatory", 5),
        ("Internal HSE System Audit", "Internal", 12),
        ("Contractor HSE Audit - AUMS", "Internal", 9),
        ("Tailings Facility Audit", "External", 4),
        ("Emergency Preparedness Audit", "Internal", 7),
        ("Explosives Magazine Audit", "Regulatory", 3),
    ]])

    drills = pd.DataFrame([{
        "Scenario": s, "Date": today - dt.timedelta(days=int(rng.integers(10, 360))),
        "Location": rng.choice(AREAS), "Participants": int(rng.integers(15, 120)),
        "Duration_min": int(rng.integers(30, 150)),
        "Effectiveness": round(float(rng.uniform(0.68, 0.96)), 3),
        "Next_Due": "=EDATE([@Date],6)",
        "Status": ('=IF([@Next_Due]<TODAY(),"Overdue",'
                   'IF([@Next_Due]-TODAY()<=60,"Due Soon","Current"))'),
    } for s in [
        "Cyanide Spill Response", "Pit Fire Evacuation", "Medical Emergency (Mass Casualty)",
        "Tailings Dam Breach", "Fuel Farm Fire", "Confined Space Rescue",
        "Hazmat Transport Incident", "Security Lockdown", "Mine Rescue Exercise"]])

    ppe = pd.DataFrame([{
        "Category": cat, "Department": dep, "Issued": iss, "Required": req,
        "Compliance": "=IFERROR([@Issued]/[@Required],0)",
        "Last_Inspection": today - dt.timedelta(days=int(rng.integers(5, 90))),
        "Status": ('=IF([@Compliance]>=0.95,"Compliant",'
                   'IF([@Compliance]>=0.85,"Monitor","Action"))'),
    } for cat, dep, iss, req in [
        ("Hard Hats", "Mining", 280, 285), ("Safety Boots", "Mining", 270, 285),
        ("Hearing Protection", "Processing", 130, 140), ("Respirators (Cyanide)", "Processing", 95, 100),
        ("Cut-resistant Gloves", "Engineering", 110, 130), ("Hi-Vis Clothing", "Mining", 285, 285),
        ("Fall Arrest Harness", "Engineering", 42, 50), ("Chemical Suits", "Laboratory", 38, 40),
        ("Eye Protection", "All", 520, 540), ("Gas Detectors", "Processing", 28, 35)]])

    equipment = pd.DataFrame([{
        "Asset_ID": f"EQP-{i:03d}", "Asset": a, "Type": t, "Location": rng.choice(AREAS),
        "Last_Inspection": today - dt.timedelta(days=int(rng.integers(20, 200))),
        "Next_Inspection": today + dt.timedelta(days=int(d)),
        "Days_To_Due": "=[@Next_Inspection]-TODAY()",
        "Status": ('=IF([@Next_Inspection]<TODAY(),"Overdue",'
                   'IF([@Next_Inspection]-TODAY()<=30,"Due Soon","Compliant"))'),
    } for i, (a, t, d) in enumerate([
        ("Fire Pump House 1", "Fire", 25), ("Foam System - Fuel Farm", "Fire", -5),
        ("SCBA Set A", "Emergency", 40), ("Overhead Crane - Workshop", "Lifting", 90),
        ("Mobile Crane 50t", "Lifting", 15), ("Gas Detection Network", "Monitoring", 60),
        ("Eyewash Stations (CIL)", "Emergency", -2), ("Emergency Generator", "Power", 120),
        ("Ambulance 1", "Medical", 30), ("Tailings Piezometers", "Geotech", 75),
        ("AED Units", "Medical", 50), ("Spill Response Trailer", "Environmental", 100)], start=1)])

    return {
        "months": months, "incidents": incidents, "activity": activity,
        "actions": actions, "risks": risks, "compliance": compliance,
        "environmental": environmental, "contractors": contractors,
        "permits": permits, "audits": audits, "drills": drills, "ppe": ppe,
        "equipment": equipment,
    }


# =============================================================================
# 8. SHEET BUILDERS
# =============================================================================

def build_lists(wb, data):
    """Hidden-ish reference sheet holding every dropdown list + named ranges."""
    ws = wb["Lists"]
    ws.sheet_view.showGridLines = False
    title_band(ws, "Reference Lists & Validation Sources", COMPANY_SUB, "R")
    ws["A5"] = ("These lists feed the data-validation dropdowns and named ranges. "
                "Add new options at the bottom of a column to extend a dropdown.")
    ws["A5"].font = Font(italic=True, color=GREY)

    lists = {
        "Type": INCIDENT_TYPES, "Class": INCIDENT_CLASSES, "IncStatus": INCIDENT_STATUS,
        "Area": AREAS, "Dept": DEPARTMENTS, "Severity": [1, 2, 3, 4, 5],
        "ActionStatus": ACTION_STATUS, "Month": ["All"] + [calendar.month_abbr[m] for m in range(1, 13)],
        "Year": ["All"] + sorted({mo["year"] for mo in data["months"]}),
        "Company": COMPANIES, "YesNo": ["Yes", "No"], "Adequacy": ADEQUACY,
        "Priority": PRIORITY, "AuditType": AUDIT_TYPES,
        "AreaFilter": ["All"] + AREAS, "DeptFilter": ["All"] + DEPARTMENTS,
    }
    ranges = {}
    start_row = 7
    for j, (key, vals) in enumerate(lists.items()):
        col = 1 + j
        h = ws.cell(start_row, col, key)
        h.font = Font(bold=True, color="FFFFFF")
        h.fill = fill(NAVY2)
        h.alignment = CENTER
        for i, v in enumerate(vals):
            ws.cell(start_row + 1 + i, col, v).alignment = CENTER
        first = coord(start_row + 1, col)
        last = coord(start_row + len(vals), col)
        ranges[key] = f"Lists!${get_column_letter(col)}${start_row + 1}:${get_column_letter(col)}${start_row + len(vals)}"
        add_defined_name(wb, f"{key}List", ranges[key])
        ws.column_dimensions[get_column_letter(col)].width = 16
    return ranges


def build_settings(wb):
    ws = wb["Settings"]
    ws.sheet_view.showGridLines = False
    title_band(ws, "Settings, Rate Bases, Targets & Thresholds", COMPANY_SUB, "E")
    ws["A5"] = ("EDIT THESE to re-tune the whole workbook. Every KPI formula references "
                "the named cells below (e.g. RATE_BASE, TARGET_TRIFR) -- never a hard-coded number.")
    ws["A5"].font = Font(italic=True, color=GREY)
    style_range(ws, "A5:E5", align=LEFT)

    for cell, txt in (("A7", "Parameter"), ("B7", "Value"), ("C7", "Named range")):
        ws[cell] = txt
        ws[cell].font = Font(bold=True, color="FFFFFF")
        ws[cell].fill = fill(NAVY2)
        ws[cell].alignment = CENTER
    r = 8
    for name, label, value, fmt in SETTINGS:
        ws.cell(r, 1, label).alignment = LEFT
        vcell = ws.cell(r, 2, value)
        vcell.number_format = fmt
        vcell.font = Font(bold=True, color=NAVY)
        vcell.alignment = CENTER
        vcell.fill = fill(LIGHT)
        vcell.border = BORDER_THIN
        ws.cell(r, 3, name).font = Font(italic=True, color=GREY, size=9)
        add_defined_name(wb, name, f"Settings!$B${r}")
        r += 1
    set_widths(ws, {"A": 42, "B": 16, "C": 22, "D": 4, "E": 10})
    ws.freeze_panes = "A8"


def build_incident_register(wb, data, list_ranges):
    """The data spine + the 'Incidents' register module (Excel Table, CF, filters)."""
    ws = wb["Incident_Register"]
    ws.sheet_view.showGridLines = False
    title_band(ws, "Incident Register  (DATA SPINE -- one row per event)",
               "Edit / append here. Year, Month, Recordable, LTI & CAR flags auto-calculate.", "T")

    nf = {"Date": FMT_DATE, "Reported": FMT_DATE, "CAR_Due": FMT_DATE,
          "Severity": "0", "Recordable": "0", "LTI": "0", "Month": "0",
          "LTI_Date": FMT_DATE, "CAR_Overdue": "0", "Days_Overdue": "0"}
    lay = write_table(ws, data["incidents"], top_row=5, left_col=1,
                      table_name="tblIncidents", number_formats=nf)
    ws.freeze_panes = "A6"

    # data-validation dropdowns on editable columns (rows 6..1000 for growth)
    dv_specs = {"Area": "AreaList", "Department": "DeptList", "Company": "CompanyList",
                "Type": "TypeList", "Class": "ClassList", "Severity": "SeverityList",
                "Status": "IncStatusList"}
    for header, listkey in dv_specs.items():
        dv = DataValidation(type="list", formula1=list_ranges[
            {"AreaList": "Area", "DeptList": "Dept", "CompanyList": "Company",
             "TypeList": "Type", "ClassList": "Class", "SeverityList": "Severity",
             "IncStatusList": "IncStatus"}[listkey]], allow_blank=True)
        ws.add_data_validation(dv)
        cl = lay.col_letter(header)
        dv.add(f"{cl}{lay.first_data_row}:{cl}1000")

    # conditional formatting: severity scale
    sev = lay.data_range("Severity", buffer=400)
    ws.conditional_formatting.add(sev, CellIsRule(operator="between", formula=["4", "5"],
                                  fill=fill(RED_BG), font=Font(bold=True, color=RED_TX)))
    ws.conditional_formatting.add(sev, CellIsRule(operator="equal", formula=["3"],
                                  fill=fill(AMBER_BG), font=Font(bold=True, color=AMBER_TX)))
    ws.conditional_formatting.add(sev, CellIsRule(operator="between", formula=["1", "2"],
                                  fill=fill(GREEN_BG), font=Font(color=GREEN_TX)))
    # overdue CAR -> highlight CAR_Due cell red
    car_col = lay.col_letter("CAR_Due")
    ov_col = lay.col_letter("CAR_Overdue")
    rng_car = f"{car_col}{lay.first_data_row}:{car_col}1000"
    ws.conditional_formatting.add(rng_car, FormulaRule(
        formula=[f"${ov_col}{lay.first_data_row}=1"], fill=fill(RED_BG),
        font=Font(bold=True, color=RED_TX)))
    # status colours
    status_cf(ws, f"{lay.col_letter('Status')}{lay.first_data_row}:"
                  f"{lay.col_letter('Status')}1000", STATUS_MAPS["incident"])

    widths = {"A": 10, "B": 12, "C": 8, "D": 10, "E": 8, "F": 22, "G": 14, "H": 22,
              "I": 16, "J": 20, "K": 9, "L": 18, "M": 12, "N": 12, "O": 12, "P": 11,
              "Q": 6, "R": 12, "S": 12, "T": 12}
    set_widths(ws, widths)


def build_activity_log(wb, data, list_ranges):
    ws = wb["Activity_Log"]
    ws.sheet_view.showGridLines = False
    title_band(ws, "Activity Log  (DATA SPINE -- per area, per month)",
               "Man-hours, leading indicators & training. Year/Month auto-calculate from Period.", "U")
    nf = {"Period": FMT_MMM, "Month": "0", "ManHours": FMT_INT, "InspScore": FMT_PCT1,
          "PPE": FMT_PCT0, "InspCompletion": FMT_PCT0, "ObsCloseRate": FMT_PCT0}
    lay = write_table(ws, data["activity"], top_row=5, left_col=1,
                      table_name="tblActivity", number_formats=nf, style="TableStyleMedium9")
    ws.freeze_panes = "A6"

    for header, listkey in {"Area": "Area", "Department": "Dept", "Company": "Company"}.items():
        dv = DataValidation(type="list", formula1=list_ranges[listkey], allow_blank=True)
        ws.add_data_validation(dv)
        cl = lay.col_letter(header)
        dv.add(f"{cl}{lay.first_data_row}:{cl}2000")

    # data bars on inspection score & PPE
    ws.conditional_formatting.add(lay.data_range("InspScore", 400),
        ColorScaleRule(start_type="num", start_value=0.7, start_color=RED_BG,
                       mid_type="num", mid_value=0.9, mid_color=AMBER_BG,
                       end_type="num", end_value=1.0, end_color=GREEN_BG))
    set_widths(ws, {"A": 10, "B": 8, "C": 10, "D": 7, "E": 22, "F": 14, "G": 22,
                    "H": 11, "I": 10, "J": 9, "K": 10, "L": 10, "M": 11, "N": 10,
                    "O": 10, "P": 8, "Q": 9, "R": 12, "S": 13, "T": 7, "U": 12})


def build_calc(wb, data):
    """Filter-aware calculation engine that feeds the Dashboard charts."""
    ws = wb["Calc"]
    ws.sheet_view.showGridLines = False
    title_band(ws, "Calculation Engine  (chart sources -- filter-aware)",
               "Auto-generated. Do not edit; everything here recalculates from the data spine.", "Z")
    months = data["months"]
    n = len(months)

    dl_inc = pred("tblIncidents", year=False, month=False)        # dept+loc only
    dl_act = pred("tblActivity", year=False, month=False)
    full_inc = pred("tblIncidents")
    ymd_inc = pred("tblIncidents", loc=False)                     # year+month+dept
    dl_actions = pred("tblActions", year=False, month=False)

    # ---- monthly timeline block (cols A..L, headers row 5) -----------------
    hdr = ["Label", "Year", "MonthName", "Total", "Recordable", "NearMiss",
           "InspScore", "ManHours", "TRIFR", "InspDone", "TrainComp", "ObsRaised",
           "", "Type", "TypeCount", "", "Area", "AreaCount", "", "Status", "StatCount",
           "", "InspPlanned", "InspComp%", "TrainAssigned", "TrainComp%"]
    for j, h in enumerate(hdr):
        c = ws.cell(5, 1 + j, h)
        if h:
            c.font = Font(bold=True, color="FFFFFF", size=9)
            c.fill = fill(NAVY2)
            c.alignment = CENTER
    base = 6
    for i, mo in enumerate(months):
        r = base + i
        ws.cell(r, 2, mo["year"])
        ws.cell(r, 3, mo["abbr"])
        ws.cell(r, 1, f'=C{r}&"-"&RIGHT(B{r},2)')
        ym_i = f"(tblIncidents[Year]=$B{r})*(tblIncidents[MonthName]=$C{r})"
        ym_a = f"(tblActivity[Year]=$B{r})*(tblActivity[MonthName]=$C{r})"
        ws.cell(r, 4, f"=SUMPRODUCT({dl_inc}*{ym_i})")
        ws.cell(r, 5, f"=SUMPRODUCT({dl_inc}*{ym_i}*tblIncidents[Recordable])")
        ws.cell(r, 6, f"=SUMPRODUCT({dl_act}*{ym_a}*tblActivity[NearMisses])")
        ws.cell(r, 7, (f"=IFERROR(SUMPRODUCT({dl_act}*{ym_a}*tblActivity[InspDone]*"
                       f"tblActivity[InspScore])/SUMPRODUCT({dl_act}*{ym_a}*"
                       f'tblActivity[InspDone]),"")')).number_format = FMT_PCT0
        ws.cell(r, 8, f"=SUMPRODUCT({dl_act}*{ym_a}*tblActivity[ManHours])")
        ws.cell(r, 9, f"=IF(H{r}=0,0,E{r}/H{r}*RATE_BASE)").number_format = FMT_DEC2
        ws.cell(r, 10, f"=SUMPRODUCT({dl_act}*{ym_a}*tblActivity[InspDone])")
        ws.cell(r, 11, f"=SUMPRODUCT({dl_act}*{ym_a}*tblActivity[TrainCompleted])")
        ws.cell(r, 12, f"=SUMPRODUCT({dl_act}*{ym_a}*tblActivity[ObsRaised])")
        ws.cell(r, 23, f"=SUMPRODUCT({dl_act}*{ym_a}*tblActivity[InspPlanned])")
        ws.cell(r, 24, f"=IFERROR(J{r}/W{r},0)").number_format = FMT_PCT0
        ws.cell(r, 25, f"=SUMPRODUCT({dl_act}*{ym_a}*tblActivity[TrainAssigned])")
        ws.cell(r, 26, f"=IFERROR(K{r}/Y{r},0)").number_format = FMT_PCT0

    # ---- by-Type block (N15..O cols 14-15) ---------------------------------
    for i, t in enumerate(INCIDENT_TYPES):
        r = base + i
        ws.cell(r, 14, t)
        ws.cell(r, 15, f'=SUMPRODUCT({full_inc}*(tblIncidents[Type]=$N{r}))')

    # ---- by-Area block (cols 17-18) ----------------------------------------
    for i, a in enumerate(AREAS):
        r = base + i
        ws.cell(r, 17, a)
        ws.cell(r, 18, f'=SUMPRODUCT({ymd_inc}*(tblIncidents[Area]=$Q{r}))')

    # ---- actions-by-status block (cols 20-21) ------------------------------
    for i, s in enumerate(ACTION_STATUS):
        r = base + i
        ws.cell(r, 20, s)
        ws.cell(r, 21, f'=SUMPRODUCT({dl_actions}*(tblActions[Status]=$T{r}))')

    set_widths(ws, {get_column_letter(c): 11 for c in range(1, 27)})
    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["N"].width = 18
    ws.column_dimensions["Q"].width = 24
    ws.column_dimensions["T"].width = 13

    return {
        "n": n, "base": base, "last": base + n - 1,
        "labels": Reference(ws, min_col=1, min_row=base, max_row=base + n - 1),
        "total": Reference(ws, min_col=4, min_row=5, max_row=base + n - 1),
        "recordable": Reference(ws, min_col=5, min_row=5, max_row=base + n - 1),
        "nearmiss": Reference(ws, min_col=6, min_row=5, max_row=base + n - 1),
        "inspscore": Reference(ws, min_col=7, min_row=5, max_row=base + n - 1),
        "trifr": Reference(ws, min_col=9, min_row=5, max_row=base + n - 1),
        "inspdone": Reference(ws, min_col=10, min_row=5, max_row=base + n - 1),
        "traincomp": Reference(ws, min_col=11, min_row=5, max_row=base + n - 1),
        "inspcomp": Reference(ws, min_col=24, min_row=5, max_row=base + n - 1),
        "type_lbl": Reference(ws, min_col=14, min_row=base, max_row=base + len(INCIDENT_TYPES) - 1),
        "type_cnt": Reference(ws, min_col=15, min_row=5, max_row=base + len(INCIDENT_TYPES) - 1),
        "area_lbl": Reference(ws, min_col=17, min_row=base, max_row=base + len(AREAS) - 1),
        "area_cnt": Reference(ws, min_col=18, min_row=5, max_row=base + len(AREAS) - 1),
        "stat_lbl": Reference(ws, min_col=20, min_row=base, max_row=base + len(ACTION_STATUS) - 1),
        "stat_cnt": Reference(ws, min_col=21, min_row=5, max_row=base + len(ACTION_STATUS) - 1),
        "ws": ws,
    }


def build_dashboard(wb, data, list_ranges, calc):
    ws = wb["Dashboard"]
    ws.sheet_view.showGridLines = False
    title_band(ws, "HSE EXECUTIVE DASHBOARD",
               "Live scorecard -- select filters below; every card & chart recalculates.", "Q")

    # ---- filter panel (row 5 labels, row 6 inputs) -------------------------
    section_header(ws, "B5:P5", "INTERACTIVE FILTERS   (choose 'All' to clear)")
    filt = [("C", "YEAR", "B", "YearList"), ("F", "MONTH", "E", "MonthList"),
            ("I", "DEPARTMENT", "H", "DeptFilterList"), ("L", "LOCATION / AREA", "K", "AreaFilterList")]
    listkey_map = {"YearList": "Year", "MonthList": "Month",
                   "DeptFilterList": "DeptFilter", "AreaFilterList": "AreaFilter"}
    for in_col, label, lab_col, listkey in filt:
        lc = ws[f"{lab_col}6"]
        lc.value = label
        lc.font = Font(bold=True, color=NAVY, size=9)
        lc.alignment = RIGHT
        cell = ws[f"{in_col}6"]
        cell.value = "All"
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill(GOLD)
        cell.alignment = CENTER
        cell.border = BORDER_BOX
        ws.merge_cells(f"{in_col}6:{get_column_letter(openpyxl.utils.column_index_from_string(in_col)+1)}6")
        dv = DataValidation(type="list", formula1=list_ranges[listkey_map[listkey]], allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(f"{in_col}6")
    ws[f"N6"] = "Report date:"
    ws["N6"].font = Font(bold=True, color=GREY, size=9)
    ws["N6"].alignment = RIGHT
    ws["O6"] = "=TODAY()"
    ws["O6"].number_format = FMT_DATE
    ws["O6"].font = Font(bold=True, color=NAVY)
    ws.merge_cells("O6:P6")

    # ---- KPI cards ----------------------------------------------------------
    pi, pa = pred("tblIncidents"), pred("tblActivity")
    pact = pred("tblActions", year=False, month=False)
    mh = f"SUMPRODUCT({pa}*tblActivity[ManHours])"
    cards = [
        # (col,row,title,formula,fmt,mode,target,status_text)
        (2, 8, "TOTAL INCIDENTS", f"=SUMPRODUCT({pi})", FMT_INT, None, None, '="period total"'),
        (6, 8, "NEAR MISSES", f"=SUMPRODUCT({pa}*tblActivity[NearMisses])", FMT_INT,
         "higher_better", "TARGET_NEARMISS", '=IF(F9>=TARGET_NEARMISS,"ON TARGET","BELOW TARGET")'),
        (10, 8, "TRIFR", (f"=IF({mh}=0,0,SUMPRODUCT({pi}*tblIncidents[Recordable])/"
                          f"{mh}*RATE_BASE)"), FMT_DEC2, "lower_better", "TARGET_TRIFR",
         '=IF(J9<=TARGET_TRIFR,"ON TARGET","ABOVE TARGET")'),
        (14, 8, "LTIFR", (f"=IF({mh}=0,0,SUMPRODUCT({pi}*tblIncidents[LTI])/"
                          f"{mh}*RATE_BASE)"), FMT_DEC2, "lower_better", "TARGET_LTIFR",
         '=IF(N9<=TARGET_LTIFR,"ON TARGET","ABOVE TARGET")'),
        (2, 13, "INSPECTION SCORE",
         (f"=IFERROR(SUMPRODUCT({pa}*tblActivity[InspDone]*tblActivity[InspScore])/"
          f"SUMPRODUCT({pa}*tblActivity[InspDone]),0)"), FMT_PCT1, "higher_better",
         "TARGET_INSPECTION", '=IF(B14>=TARGET_INSPECTION,"ON TARGET","BELOW TARGET")'),
        (6, 13, "TRAINING COMPLETION",
         (f"=IFERROR(SUMPRODUCT({pa}*tblActivity[TrainCompleted])/"
          f"SUMPRODUCT({pa}*tblActivity[TrainAssigned]),0)"), FMT_PCT1, "higher_better",
         "TARGET_TRAINING", '=IF(F14>=TARGET_TRAINING,"ON TARGET","BELOW TARGET")'),
        (10, 13, "ENV COMPLIANCE", "=ENV_COMPLIANCE_PCT", FMT_PCT1, "higher_better",
         "TARGET_ENV", '=IF(J14>=TARGET_ENV,"ON TARGET","BELOW TARGET")'),
        (14, 13, "DAYS SINCE LAST LTI",
         '=IF(MAX(tblIncidents[LTI_Date])=0,"N/A",TODAY()-MAX(tblIncidents[LTI_Date]))',
         FMT_INT, "days_lti", None, '="days incident-free"'),
        (2, 18, "OPEN ACTIONS", f'=SUMPRODUCT({pact}*(tblActions[Status]<>"Closed"))',
         FMT_INT, None, None, '="awaiting close-out"'),
        (6, 18, "OVERDUE ACTIONS", f"=SUMPRODUCT({pact}*tblActions[Overdue])", FMT_INT,
         "zero_best", None, '=IF(F19=0,"NONE OVERDUE","ACTION REQUIRED")'),
        (10, 18, "RECORDABLES", f"=SUMPRODUCT({pi}*tblIncidents[Recordable])", FMT_INT,
         None, None, '="MTC+RWC+LTI"'),
        (14, 18, "MAN-HOURS (PERIOD)", f"=SUMPRODUCT({pa}*tblActivity[ManHours])", FMT_INT,
         None, None, '="exposure hours"'),
    ]
    for col, row, title, formula, fmt, mode, target, stext in cards:
        accent = GOLD if mode is None else NAVY2
        vaddr, saddr = kpi_card(ws, col, row, title=title, formula=formula, fmt=fmt,
                                accent=accent, suffix_formula=stext)
        if mode:
            traffic_light(ws, vaddr, saddr, mode, target)
        else:
            ws[saddr].fill = fill(NAVY2)

    # ---- charts -------------------------------------------------------------
    cws = calc["ws"]
    section_header(ws, "B23:P23", "TRENDS & ANALYTICS")
    line_chart(cws, "Incident Trend (multi-year, monthly)", calc["labels"],
               [calc["total"], calc["recordable"]], "B25", ws, y_title="count")
    doughnut_chart(cws, "Incidents by Type", calc["type_lbl"], calc["type_cnt"], "J25", ws)
    bar_chart(cws, "Incidents by Location (high-risk ranking)", calc["area_lbl"],
              calc["area_cnt"], "B41", ws, bar_dir="bar", colour=GOLD)
    line_chart(cws, "Inspection Score Trend", calc["labels"], [calc["inspscore"]],
               "J41", ws, y_title="%", colours=["2E8B57"])
    combo_chart(cws, "Leading vs Lagging Indicators", calc["labels"],
                Reference(cws, min_col=6, min_row=5, max_row=calc["last"]),   # NearMiss bars
                Reference(cws, min_col=5, min_row=5, max_row=calc["last"]),   # Recordable line
                "B57", ws, bar_title="near misses", line_title="recordables")
    bar_chart(cws, "Corrective Actions by Status", calc["stat_lbl"], calc["stat_cnt"],
              "J57", ws, colour=NAVY2, data_labels=True)

    set_widths(ws, {get_column_letter(c): 9.5 for c in range(1, 18)})
    ws.column_dimensions["A"].width = 2
    ws.freeze_panes = "A4"


def build_near_miss(wb, data, calc):
    ws = wb["Near_Miss"]
    ws.sheet_view.showGridLines = False
    title_band(ws, "Near-Miss Reporting & Ratio", COMPANY_SUB, "P")
    pa = pred("tblActivity")
    pi = pred("tblIncidents")
    # KPI cards
    kpi_card(ws, 2, 5, title="NEAR MISSES (period)", formula=f"=SUMPRODUCT({pa}*tblActivity[NearMisses])")
    kpi_card(ws, 6, 5, title="RECORDABLES (period)", formula=f"=SUMPRODUCT({pi}*tblIncidents[Recordable])")
    v, s = kpi_card(ws, 10, 5, title="NEAR-MISS : RECORDABLE",
                    formula=(f"=IFERROR(SUMPRODUCT({pa}*tblActivity[NearMisses])/"
                             f"SUMPRODUCT({pi}*tblIncidents[Recordable]),0)"),
                    fmt=FMT_DEC2, accent=NAVY2,
                    suffix_formula='=IF(J6>=NEARMISS_RATIO_TGT,"HEALTHY","IMPROVE REPORTING")')
    traffic_light(ws, v, s, "higher_better", "NEARMISS_RATIO_TGT")
    kpi_card(ws, 14, 5, title="HAZARDS REPORTED", formula=f"=SUMPRODUCT({pa}*tblActivity[Hazards])")

    section_header(ws, "B11:P11", "REPORTING TRENDS")
    cws = calc["ws"]
    combo_chart(cws, "Near-Miss vs Recordable Trend", calc["labels"],
                Reference(cws, min_col=6, min_row=5, max_row=calc["last"]),
                Reference(cws, min_col=5, min_row=5, max_row=calc["last"]),
                "B13", ws, bar_title="near misses", line_title="recordables")
    # near miss by area (built on this sheet)
    r0 = 30
    ws.cell(r0, 2, "Area").font = Font(bold=True, color="FFFFFF")
    ws.cell(r0, 3, "Near Misses").font = Font(bold=True, color="FFFFFF")
    ws.cell(r0, 2).fill = fill(NAVY2)
    ws.cell(r0, 3).fill = fill(NAVY2)
    pdl = pred("tblActivity", loc=False)
    for i, a in enumerate(AREAS):
        rr = r0 + 1 + i
        ws.cell(rr, 2, a)
        ws.cell(rr, 3, f'=SUMPRODUCT({pdl}*(tblActivity[Area]=$B{rr})*tblActivity[NearMisses])')
    bar_chart(ws, "Near Misses by Area", Reference(ws, min_col=2, min_row=r0 + 1, max_row=r0 + len(AREAS)),
              Reference(ws, min_col=3, min_row=r0, max_row=r0 + len(AREAS)), "J13", ws,
              bar_dir="bar", colour=GOLD)
    set_widths(ws, {get_column_letter(c): 10 for c in range(1, 17)})
    ws.column_dimensions["B"].width = 24


def build_risk_matrix(wb, data, list_ranges):
    ws = wb["Risk_Matrix"]
    ws.sheet_view.showGridLines = False
    title_band(ws, "Risk Matrix (5x5) & Risk Register", COMPANY_SUB, "N")

    section_header(ws, "B5:H5", "5 x 5 RISK ASSESSMENT MATRIX  (Likelihood x Consequence)")
    likelihood = ["5 Almost Certain", "4 Likely", "3 Possible", "2 Unlikely", "1 Rare"]
    consequence = ["1 Insignificant", "2 Minor", "3 Moderate", "4 Major", "5 Severe"]
    grid_top, grid_left = 7, 3        # C7
    # corner + consequence header
    cc = ws.cell(grid_top, grid_left - 1, "L  \\  C")
    cc.fill = fill(NAVY)
    cc.font = Font(bold=True, color="FFFFFF")
    cc.alignment = CENTER
    cc.border = BORDER_THIN
    for j, c in enumerate(consequence):
        cell = ws.cell(grid_top, grid_left + j, c)
        cell.fill = fill(NAVY2)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.alignment = CENTER
        cell.border = BORDER_THIN
    for i, lk in enumerate(likelihood):
        Lval = 5 - i
        rcell = ws.cell(grid_top + 1 + i, grid_left - 1, lk)
        rcell.fill = fill(NAVY2)
        rcell.font = Font(bold=True, color="FFFFFF", size=9)
        rcell.alignment = CENTER
        rcell.border = BORDER_THIN
        for j in range(5):
            Cval = j + 1
            cell = ws.cell(grid_top + 1 + i, grid_left + j, Lval * Cval)
            cell.alignment = CENTER
            cell.font = Font(bold=True, color=SLATE)
            cell.border = BORDER_THIN
    grid_range = f"{coord(grid_top + 1, grid_left)}:{coord(grid_top + 5, grid_left + 4)}"
    for op, lo, hi, colr in [("between", "1", "4", RISK_LOW), ("between", "5", "9", RISK_MED),
                             ("between", "10", "15", RISK_HIGH), ("between", "16", "25", RISK_CRIT)]:
        ws.conditional_formatting.add(grid_range, CellIsRule(
            operator=op, formula=[lo, hi], fill=fill(colr)))
    # legend
    leg = [("Low (1-4)", RISK_LOW), ("Medium (5-9)", RISK_MED),
           ("High (10-15)", RISK_HIGH), ("Critical (16-25)", RISK_CRIT)]
    for i, (txt, colr) in enumerate(leg):
        cell = ws.cell(grid_top + 1 + i, grid_left + 6, txt)
        cell.fill = fill(colr)
        cell.alignment = CENTER
        cell.border = BORDER_THIN
        cell.font = Font(bold=True, color=SLATE)

    # ---- risk register ------------------------------------------------------
    section_header(ws, "B15:M15", "RISK REGISTER  (score = Likelihood x Consequence; auto-rated)")
    nf = {"Score": "0", "Likelihood": "0", "Consequence": "0", "Review_Date": FMT_DATE}
    lay = write_table(ws, data["risks"], top_row=16, left_col=2, table_name="tblRisks",
                      number_formats=nf)
    for header, listkey in {"Likelihood": "Severity", "Consequence": "Severity",
                            "Control_Adequacy": "Adequacy"}.items():
        dv = DataValidation(type="list", formula1=list_ranges[listkey], allow_blank=True)
        ws.add_data_validation(dv)
        cl = lay.col_letter(header)
        dv.add(f"{cl}{lay.first_data_row}:{cl}200")
    # CF on score + level
    sc = lay.data_range("Score", 100)
    for lo, hi, colr in [("1", "4", RISK_LOW), ("5", "9", RISK_MED),
                         ("10", "15", RISK_HIGH), ("16", "25", RISK_CRIT)]:
        ws.conditional_formatting.add(sc, CellIsRule(operator="between", formula=[lo, hi],
                                      fill=fill(colr), font=Font(bold=True)))
    status_cf(ws, f"{lay.col_letter('Risk_Level')}{lay.first_data_row}:"
                  f"{lay.col_letter('Risk_Level')}200", STATUS_MAPS["risk"])
    status_cf(ws, f"{lay.col_letter('Control_Adequacy')}{lay.first_data_row}:"
                  f"{lay.col_letter('Control_Adequacy')}200",
              {"Adequate": GREEN_BG, "Needs Improvement": AMBER_BG, "Inadequate": RED_BG})
    set_widths(ws, {"A": 2, "B": 11, "C": 34, "D": 20, "E": 18, "F": 11, "G": 12,
                    "H": 8, "I": 11, "J": 18, "K": 14, "L": 20})


def build_inspections(wb, data, calc):
    ws = wb["Inspections"]
    ws.sheet_view.showGridLines = False
    title_band(ws, "Inspections -- Completion, Score & Target", COMPANY_SUB, "P")
    pa = pred("tblActivity")
    v, s = kpi_card(ws, 2, 5, title="INSPECTION SCORE",
                    formula=(f"=IFERROR(SUMPRODUCT({pa}*tblActivity[InspDone]*tblActivity[InspScore])/"
                             f"SUMPRODUCT({pa}*tblActivity[InspDone]),0)"), fmt=FMT_PCT1,
                    accent=NAVY2, suffix_formula='=IF(B6>=TARGET_INSPECTION,"ON TARGET","BELOW TARGET")')
    traffic_light(ws, v, s, "higher_better", "TARGET_INSPECTION")
    v2, s2 = kpi_card(ws, 6, 5, title="COMPLETION RATE",
                      formula=(f"=IFERROR(SUMPRODUCT({pa}*tblActivity[InspDone])/"
                               f"SUMPRODUCT({pa}*tblActivity[InspPlanned]),0)"), fmt=FMT_PCT1,
                      accent=NAVY2, suffix_formula='=IF(F6>=0.95,"ON TARGET","BEHIND")')
    traffic_light(ws, v2, s2, "higher_better", "TARGET_INSPECTION")
    kpi_card(ws, 10, 5, title="INSPECTIONS DONE", formula=f"=SUMPRODUCT({pa}*tblActivity[InspDone])")
    kpi_card(ws, 14, 5, title="INSPECTIONS PLANNED", formula=f"=SUMPRODUCT({pa}*tblActivity[InspPlanned])")

    section_header(ws, "B11:P11", "TREND vs TARGET")
    cws = calc["ws"]
    combo_chart(cws, "Inspection Completion % (bars) vs Score % (line)", calc["labels"],
                Reference(cws, min_col=24, min_row=5, max_row=calc["last"]),
                Reference(cws, min_col=7, min_row=5, max_row=calc["last"]),
                "B13", ws, bar_title="completion", line_title="score")
    line_chart(cws, "Inspections Completed (count)", calc["labels"], [calc["inspdone"]],
               "J13", ws, colours=[GOLD])
    set_widths(ws, {get_column_letter(c): 10 for c in range(1, 17)})


def build_training(wb, data):
    ws = wb["Training"]
    ws.sheet_view.showGridLines = False
    title_band(ws, "Training Compliance", COMPANY_SUB, "P")
    pa = pred("tblActivity")
    v, s = kpi_card(ws, 2, 5, title="TRAINING COMPLETION",
                    formula=(f"=IFERROR(SUMPRODUCT({pa}*tblActivity[TrainCompleted])/"
                             f"SUMPRODUCT({pa}*tblActivity[TrainAssigned]),0)"), fmt=FMT_PCT1,
                    accent=NAVY2, suffix_formula='=IF(B6>=TARGET_TRAINING,"ON TARGET","BELOW TARGET")')
    traffic_light(ws, v, s, "higher_better", "TARGET_TRAINING")
    kpi_card(ws, 6, 5, title="COMPLETED", formula=f"=SUMPRODUCT({pa}*tblActivity[TrainCompleted])")
    kpi_card(ws, 10, 5, title="ASSIGNED", formula=f"=SUMPRODUCT({pa}*tblActivity[TrainAssigned])")
    kpi_card(ws, 14, 5, title="OUTSTANDING",
             formula=(f"=SUMPRODUCT({pa}*tblActivity[TrainAssigned])-"
                      f"SUMPRODUCT({pa}*tblActivity[TrainCompleted])"))

    section_header(ws, "B11:P11", "COMPLIANCE BY DEPARTMENT")
    r0 = 13
    ws.cell(r0, 2, "Department").font = Font(bold=True, color="FFFFFF")
    ws.cell(r0, 3, "Completion %").font = Font(bold=True, color="FFFFFF")
    ws.cell(r0, 4, "Target").font = Font(bold=True, color="FFFFFF")
    for cc in (2, 3, 4):
        ws.cell(r0, cc).fill = fill(NAVY2)
        ws.cell(r0, cc).alignment = CENTER
    pdept = pred("tblActivity", dept=False, loc=False)   # year+month only
    for i, d in enumerate(DEPARTMENTS):
        rr = r0 + 1 + i
        ws.cell(rr, 2, d)
        ws.cell(rr, 3, (f'=IFERROR(SUMPRODUCT({pdept}*(tblActivity[Department]=$B{rr})*'
                        f'tblActivity[TrainCompleted])/SUMPRODUCT({pdept}*'
                        f'(tblActivity[Department]=$B{rr})*tblActivity[TrainAssigned]),0)')
                       ).number_format = FMT_PCT0
        ws.cell(rr, 4, "=TARGET_TRAINING").number_format = FMT_PCT0
    last = r0 + len(DEPARTMENTS)
    ws.conditional_formatting.add(f"C{r0+1}:C{last}", ColorScaleRule(
        start_type="num", start_value=0.7, start_color=RED_BG,
        mid_type="num", mid_value=0.85, mid_color=AMBER_BG,
        end_type="num", end_value=1.0, end_color=GREEN_BG))
    combo_chart(ws, "Training Completion by Department vs Target",
                Reference(ws, min_col=2, min_row=r0 + 1, max_row=last),
                Reference(ws, min_col=3, min_row=r0, max_row=last),
                Reference(ws, min_col=4, min_row=r0, max_row=last),
                "F13", ws, bar_title="completion", line_title="target")
    set_widths(ws, {"A": 2, "B": 18, "C": 14, "D": 10})
    set_widths(ws, {get_column_letter(c): 10 for c in range(5, 17)})


def build_corrective_actions(wb, data, list_ranges, calc):
    ws = wb["Corrective_Actions"]
    ws.sheet_view.showGridLines = False
    title_band(ws, "Corrective Actions Tracker", COMPANY_SUB, "M")
    pact = pred("tblActions", year=False, month=False)
    kpi_card(ws, 2, 5, title="OPEN ACTIONS", formula=f'=SUMPRODUCT({pact}*(tblActions[Status]<>"Closed"))')
    v, s = kpi_card(ws, 6, 5, title="OVERDUE", formula=f"=SUMPRODUCT({pact}*tblActions[Overdue])",
                    accent=NAVY2, suffix_formula='=IF(F6=0,"NONE OVERDUE","ACTION REQUIRED")')
    traffic_light(ws, v, s, "zero_best")
    kpi_card(ws, 10, 5, title="CLOSED", formula=f'=SUMPRODUCT({pact}*(tblActions[Status]="Closed"))')
    kpi_card(ws, 14, 5, title="DUE SOON", formula=f'=SUMPRODUCT({pact}*(tblActions[Status]="Due Soon"))')

    section_header(ws, "B11:M11", "ACTION REGISTER")
    nf = {"Raised": FMT_DATE, "Due": FMT_DATE, "Overdue": "0", "Days_Overdue": "0"}
    lay = write_table(ws, data["actions"], top_row=12, left_col=2, table_name="tblActions",
                      number_formats=nf)
    ws.freeze_panes = "A13"
    dv = DataValidation(type="list", formula1=list_ranges["ActionStatus"], allow_blank=True)
    ws.add_data_validation(dv)
    sc = lay.col_letter("Status")
    dv.add(f"{sc}{lay.first_data_row}:{sc}1000")
    status_cf(ws, f"{sc}{lay.first_data_row}:{sc}1000", STATUS_MAPS["action"])
    od = lay.col_letter("Days_Overdue")
    ws.conditional_formatting.add(f"{od}{lay.first_data_row}:{od}1000",
        CellIsRule(operator="greaterThan", formula=["0"], fill=fill(RED_BG),
                   font=Font(bold=True, color=RED_TX)))
    # actions-by-status doughnut (from Calc)
    cws = calc["ws"]
    doughnut_chart(cws, "Actions by Status", calc["stat_lbl"], calc["stat_cnt"], "I5", ws)
    set_widths(ws, {"A": 2, "B": 10, "C": 14, "D": 30, "E": 12, "F": 12, "G": 12,
                    "H": 14, "I": 18, "J": 9, "K": 10, "L": 9, "M": 11})


def build_compliance(wb, data, list_ranges):
    ws = wb["Compliance"]
    ws.sheet_view.showGridLines = False
    title_band(ws, "Regulatory Compliance (Ghana)", COMPANY_SUB, "K")
    v, s = kpi_card(ws, 2, 5, title="OVERALL COMPLIANCE",
                    formula='=IFERROR(COUNTIF(tblCompliance[Status],"Compliant")/ROWS(tblCompliance),0)',
                    fmt=FMT_PCT0, accent=NAVY2,
                    suffix_formula='=IF(B6>=0.95,"COMPLIANT","ATTENTION")')
    traffic_light(ws, v, s, "higher_better", "0.95")
    kpi_card(ws, 6, 5, title="DUE SOON", formula='=COUNTIF(tblCompliance[Status],"Due Soon")')
    v2, s2 = kpi_card(ws, 10, 5, title="OVERDUE", formula='=COUNTIF(tblCompliance[Status],"Overdue")',
                      accent=NAVY2, suffix_formula='=IF(J6=0,"NONE OVERDUE","ACTION REQUIRED")')
    traffic_light(ws, v2, s2, "zero_best")

    section_header(ws, "B11:J11", "REGULATORY REGISTER  (status auto-derived from Due date vs TODAY)")
    nf = {"Frequency_Months": "0", "Last_Completed": FMT_DATE, "Due_Date": FMT_DATE,
          "Days_To_Due": "0"}
    lay = write_table(ws, data["compliance"], top_row=12, left_col=2,
                      table_name="tblCompliance", number_formats=nf)
    ws.freeze_panes = "A13"
    status_cf(ws, f"{lay.col_letter('Status')}{lay.first_data_row}:"
                  f"{lay.col_letter('Status')}200", STATUS_MAPS["compliance"])
    # status doughnut built on this sheet
    r0 = 5
    ws.cell(r0, 13, "Status").font = Font(bold=True, color="FFFFFF")
    ws.cell(r0, 14, "Count").font = Font(bold=True, color="FFFFFF")
    ws.cell(r0, 13).fill = fill(NAVY2)
    ws.cell(r0, 14).fill = fill(NAVY2)
    for i, st in enumerate(["Compliant", "Due Soon", "Overdue"]):
        ws.cell(r0 + 1 + i, 13, st)
        ws.cell(r0 + 1 + i, 14, f'=COUNTIF(tblCompliance[Status],$M{r0+1+i})')
    doughnut_chart(ws, "Compliance Status", Reference(ws, min_col=13, min_row=r0 + 1, max_row=r0 + 3),
                   Reference(ws, min_col=14, min_row=r0, max_row=r0 + 3), "P5", ws)
    set_widths(ws, {"A": 2, "B": 34, "C": 24, "D": 12, "E": 14, "F": 14, "G": 14,
                    "H": 14, "I": 14, "J": 11, "K": 3, "M": 12, "N": 8})


def build_environmental(wb, data):
    ws = wb["Environmental"]
    ws.sheet_view.showGridLines = False
    title_band(ws, "Environmental Monitoring & Limits", COMPANY_SUB, "T")
    # ENV_COMPLIANCE_PCT lives at C4 (named range created up-front)
    ws["B4"] = "Environmental compliance % (readings within limit, filter-aware):"
    ws["B4"].font = Font(bold=True, color=NAVY)
    ws["B4"].alignment = RIGHT
    ws["C4"] = ('=IFERROR(SUMPRODUCT(((F_Year="All")+(F_Year="")+(tblEnv[Year]=F_Year))*'
                '((F_Month="All")+(F_Month="")+(tblEnv[MonthName]=F_Month))*'
                '(tblEnv[PM10_OK]+tblEnv[pH_OK]+tblEnv[CN_OK]))/'
                '(3*SUMPRODUCT(((F_Year="All")+(F_Year="")+(tblEnv[Year]=F_Year))*'
                '((F_Month="All")+(F_Month="")+(tblEnv[MonthName]=F_Month)))),1)')
    ws["C4"].number_format = FMT_PCT1
    ws["C4"].font = Font(bold=True, size=14, color=NAVY)
    ws["C4"].fill = fill(LIGHT)
    ws["C4"].border = BORDER_BOX

    section_header(ws, "B6:T6", "MONTHLY MONITORING  (regulated parameters flagged vs limit)")
    nf = {"Period": FMT_MMM, "Waste_t": "#,##0.0", "Recycling": FMT_PCT0,
          "Energy_MWh": FMT_INT, "Water_m3": FMT_INT, "Fuel_L": FMT_INT,
          "PM10": "0.0", "pH": "0.00", "WAD_CN": "0.0", "PM10_Limit": "0",
          "CN_Limit": "0", "PM10_OK": "0", "pH_OK": "0", "CN_OK": "0"}
    lay = write_table(ws, data["environmental"], top_row=7, left_col=2, table_name="tblEnv",
                      number_formats=nf, style="TableStyleMedium7")
    ws.freeze_panes = "A8"
    # exceedance highlighting on the OK flag columns (0 = exceedance)
    for flagcol in ("PM10_OK", "pH_OK", "CN_OK"):
        cl = lay.col_letter(flagcol)
        rng = f"{cl}{lay.first_data_row}:{cl}{lay.last_data_row}"
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=["0"],
                                      fill=fill(RED_BG), font=Font(bold=True, color=RED_TX)))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=["1"],
                                      fill=fill(GREEN_BG), font=Font(color=GREEN_TX)))
    # trend arrows (UP/DOWN/FLAT) colour
    for tcol, up_is_bad in (("Energy_Trend", True), ("Water_Trend", True)):
        cl = lay.col_letter(tcol)
        rng = f"{cl}{lay.first_data_row}:{cl}{lay.last_data_row}"
        bad, good = (RED_BG, GREEN_BG) if up_is_bad else (GREEN_BG, RED_BG)
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"UP"'],
                                      fill=fill(bad), font=Font(bold=True, color=RED_TX)))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"DOWN"'],
                                      fill=fill(good), font=Font(bold=True, color=GREEN_TX)))

    # charts vs regulatory limits
    cats = Reference(ws, min_col=lay.col("Period"), min_row=lay.first_data_row, max_row=lay.last_data_row)
    pm10 = Reference(ws, min_col=lay.col("PM10"), min_row=lay.top_row, max_row=lay.last_data_row)
    pm10L = Reference(ws, min_col=lay.col("PM10_Limit"), min_row=lay.top_row, max_row=lay.last_data_row)
    cn = Reference(ws, min_col=lay.col("WAD_CN"), min_row=lay.top_row, max_row=lay.last_data_row)
    cnL = Reference(ws, min_col=lay.col("CN_Limit"), min_row=lay.top_row, max_row=lay.last_data_row)
    line_chart(ws, "PM10 Dust vs EPA Limit", cats, [pm10, pm10L], "B33", ws,
               y_title="ug/m3", colours=[GOLD, "C0504D"])
    line_chart(ws, "WAD Cyanide vs ICMC Limit", cats, [cn, cnL], "J33", ws,
               y_title="mg/L", colours=["2E8B57", "C0504D"])
    energy = Reference(ws, min_col=lay.col("Energy_MWh"), min_row=lay.top_row, max_row=lay.last_data_row)
    bar_chart(ws, "Energy Consumption (MWh)", cats, energy, "B49", ws, colour=NAVY2)
    water = Reference(ws, min_col=lay.col("Water_m3"), min_row=lay.top_row, max_row=lay.last_data_row)
    bar_chart(ws, "Water Consumption (m3)", cats, water, "J49", ws, colour="4BACC6")
    set_widths(ws, {"A": 2, "B": 10, "C": 8, "D": 10, "E": 10, "F": 11, "G": 11, "H": 11,
                    "I": 9, "J": 8, "K": 9, "L": 11, "M": 10, "N": 9, "O": 8, "P": 8,
                    "Q": 12, "R": 12})


def build_contractors(wb, data):
    ws = wb["Contractors"]
    ws.sheet_view.showGridLines = False
    title_band(ws, "Contractor HSE Performance", COMPANY_SUB, "I")
    section_header(ws, "B5:I5", "OWNER vs CONTRACTOR  (man-hours & recordables live from the data spine)")
    nf = {"ManHours": FMT_INT, "Recordables": "0", "LTIs": "0", "TRIFR": FMT_DEC2,
          "LTIFR": FMT_DEC2, "Target": FMT_DEC2}
    lay = write_table(ws, data["contractors"], top_row=6, left_col=2,
                      table_name="tblContractors", number_formats=nf)
    # TRIFR conditional formatting vs target
    tcol = lay.col_letter("TRIFR")
    rng = f"{tcol}{lay.first_data_row}:{tcol}{lay.last_data_row}"
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f"${tcol}{lay.first_data_row}>TARGET_TRIFR"], fill=fill(RED_BG),
        font=Font(bold=True, color=RED_TX)))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f"${tcol}{lay.first_data_row}<=TARGET_TRIFR"], fill=fill(GREEN_BG),
        font=Font(bold=True, color=GREEN_TX)))
    combo_chart(ws, "TRIFR by Company vs Target",
                Reference(ws, min_col=lay.col("Company"), min_row=lay.first_data_row, max_row=lay.last_data_row),
                Reference(ws, min_col=lay.col("TRIFR"), min_row=lay.top_row, max_row=lay.last_data_row),
                Reference(ws, min_col=lay.col("Target"), min_row=lay.top_row, max_row=lay.last_data_row),
                "B16", ws, bar_title="TRIFR", line_title="target")
    set_widths(ws, {"A": 2, "B": 26, "C": 34, "D": 12, "E": 12, "F": 8, "G": 10, "H": 10, "I": 10})


def _register_sheet(wb, sheet, title, df, table_name, nf, status_key, status_col="Status",
                    extra_dv=None, list_ranges=None, score_cols=None):
    """Generic builder for the supporting registers (Permits/Audits/Drills/PPE/Equipment)."""
    ws = wb[sheet]
    ws.sheet_view.showGridLines = False
    last_col = get_column_letter(1 + len(df.columns) + 1)
    title_band(ws, title, COMPANY_SUB, last_col)
    section_header(ws, f"B5:{last_col}5", title.upper())
    lay = write_table(ws, df, top_row=6, left_col=2, table_name=table_name, number_formats=nf)
    ws.freeze_panes = "A7"
    if status_key:
        cl = lay.col_letter(status_col)
        status_cf(ws, f"{cl}{lay.first_data_row}:{cl}500", STATUS_MAPS[status_key])
    for col in (score_cols or []):
        cl = lay.col_letter(col)
        ws.conditional_formatting.add(f"{cl}{lay.first_data_row}:{cl}500", ColorScaleRule(
            start_type="num", start_value=0.7, start_color=RED_BG,
            mid_type="num", mid_value=0.85, mid_color=AMBER_BG,
            end_type="num", end_value=1.0, end_color=GREEN_BG))
    # reasonable widths
    set_widths(ws, {"A": 2})
    for j, h in enumerate(df.columns):
        ws.column_dimensions[get_column_letter(2 + j)].width = max(11, min(34, len(str(h)) + 8))
    return lay


def build_supporting_registers(wb, data):
    _register_sheet(wb, "Permits", "Permit-to-Work Register", data["permits"], "tblPermits",
                    {"Issue_Date": FMT_DATE, "Expiry_Date": FMT_DATE, "Days_To_Expiry": "0"},
                    "permit")
    _register_sheet(wb, "Audits", "Audit Register", data["audits"], "tblAudits",
                    {"Date": FMT_DATE, "Score": FMT_PCT0, "Findings": "0", "Closed_Findings": "0"},
                    "audit", score_cols=["Score"])
    _register_sheet(wb, "Drills", "Emergency Drills Register", data["drills"], "tblDrills",
                    {"Date": FMT_DATE, "Participants": "0", "Duration_min": "0",
                     "Effectiveness": FMT_PCT0, "Next_Due": FMT_DATE}, "drill",
                    score_cols=["Effectiveness"])
    _register_sheet(wb, "PPE", "PPE Compliance Register", data["ppe"], "tblPPE",
                    {"Issued": "0", "Required": "0", "Compliance": FMT_PCT0,
                     "Last_Inspection": FMT_DATE}, "ppe", score_cols=["Compliance"])
    _register_sheet(wb, "Equipment", "Safety-Critical Equipment Register", data["equipment"],
                    "tblEquipment", {"Last_Inspection": FMT_DATE, "Next_Inspection": FMT_DATE,
                                     "Days_To_Due": "0"}, "equip")


def build_readme(wb):
    ws = wb["ReadMe"]
    ws.sheet_view.showGridLines = False
    title_band(ws, "ReadMe -- Architecture & How To Use", COMPANY_SUB, "J")
    blocks = [
        ("HOW THIS WORKBOOK WORKS", [
            "This is a LIVE workbook, not a static report. There is one auditable data spine:",
            "   - Incident_Register  (one row per event)",
            "   - Activity_Log       (per area, per month: man-hours & leading indicators)",
            "   - Settings           (rate bases, targets & thresholds -- named cells)",
            "",
            "EVERY KPI, status light and chart on every other sheet is a native Excel FORMULA",
            "(SUMPRODUCT / COUNTIF / AVERAGEIFS / TODAY / EDATE) that points back at the spine.",
            "When you add a new row to a data sheet, the Excel Tables auto-expand and every number,",
            "traffic-light and chart recalculates automatically -- no redesign, no re-running code.",
        ]),
        ("YOU ONLY EVER EDIT TWO SHEETS", [
            "   1. Incident_Register  -- append incidents (dropdowns guide Type/Class/Status/Area...)",
            "   2. Activity_Log       -- append monthly activity per area",
            "Optionally tune Settings (targets / limits) and extend dropdowns on the Lists sheet.",
            "Columns shaded as calculated (Year, Recordable, LTI, CAR_Overdue, Status, Due_Date...)",
            "are formulas -- do not overtype them; Excel copies them down to new rows for you.",
        ]),
        ("THE INTERACTIVE DASHBOARD", [
            "The Dashboard has Year / Month / Department / Location dropdowns. Choosing a value",
            "re-drives every headline KPI (SUMPRODUCT with full 'All' handling) and all six charts.",
            "Set a filter back to 'All' to clear it.",
        ]),
        ("KEY DEFINITIONS", [
            "Recordable      = Medical Treatment + Restricted Work + Lost Time Injury",
            "TRIFR           = Recordables / man-hours x RATE_BASE (default 1,000,000)",
            "LTIFR           = Lost Time Injuries / man-hours x RATE_BASE",
            "(Fatalities are modelled as severity-5 Lost Time Injuries, so they count in both.)",
            "Near-miss ratio = near misses / recordables (a leading-indicator health check)",
            "Overdue action  = Status <> Closed AND Due date < TODAY()",
        ]),
        ("SHEETS", [
            "Dashboard, Incident_Register, Activity_Log, Settings, Lists, Calc (chart engine),",
            "Near_Miss, Risk_Matrix, Inspections, Training, Corrective_Actions, Compliance,",
            "Environmental, Contractors, Permits, Audits, Drills, PPE, Equipment.",
            "(Incident_Register doubles as the 'Incidents' register module: it is an Excel Table",
            " with autofilter and conditional formatting on severity and overdue CARs.)",
        ]),
        ("NOTE ON RECALCULATION", [
            "openpyxl writes formulas but does not compute them. The first time you OPEN the file",
            "in Excel, all formulas calculate (fullCalcOnLoad is set). If a cell shows 0 before",
            "opening in Excel, that is expected -- Excel fills it in on open.",
        ]),
    ]
    r = 5
    for header, lines in blocks:
        section_header(ws, f"B{r}:J{r}", header)
        r += 1
        for ln in lines:
            ws.cell(r, 2, ln).font = Font(color=SLATE, size=10,
                                          bold=ln.startswith("   ") is False and ln.isupper())
            ws.cell(r, 2).alignment = Alignment(horizontal="left", vertical="center")
            r += 1
        r += 1
    set_widths(ws, {"A": 2, "B": 100})


# =============================================================================
# 9. ORCHESTRATOR
# =============================================================================

def build_workbook(data) -> Workbook:
    wb = Workbook()
    # create all sheets in tab order, drop the default
    order = ["ReadMe", "Dashboard", "Incident_Register", "Activity_Log", "Settings",
             "Lists", "Near_Miss", "Risk_Matrix", "Inspections", "Training",
             "Corrective_Actions", "Compliance", "Environmental", "Contractors",
             "Permits", "Audits", "Drills", "PPE", "Equipment", "Calc"]
    wb.remove(wb.active)
    for name in order:
        wb.create_sheet(name)

    # defined names that point at fixed anchor cells (decoupled from build order)
    add_defined_name(wb, "F_Year", ADDR_FILTER_YEAR)
    add_defined_name(wb, "F_Month", ADDR_FILTER_MONTH)
    add_defined_name(wb, "F_Dept", ADDR_FILTER_DEPT)
    add_defined_name(wb, "F_Loc", ADDR_FILTER_LOC)
    add_defined_name(wb, "ENV_COMPLIANCE_PCT", ADDR_ENV_PCT)

    # build supporting / source sheets first (defines list named ranges + settings)
    list_ranges = build_lists(wb, data)
    build_settings(wb)

    # data spine
    build_incident_register(wb, data, list_ranges)
    build_activity_log(wb, data, list_ranges)

    # calculation engine + dashboard
    calc = build_calc(wb, data)
    build_dashboard(wb, data, list_ranges, calc)

    # module sheets
    build_near_miss(wb, data, calc)
    build_risk_matrix(wb, data, list_ranges)
    build_inspections(wb, data, calc)
    build_training(wb, data)
    build_corrective_actions(wb, data, list_ranges, calc)
    build_compliance(wb, data, list_ranges)
    build_environmental(wb, data)
    build_contractors(wb, data)
    build_supporting_registers(wb, data)
    build_readme(wb)

    # presentation: hide the engine sheets, force recalculation on open
    wb["Calc"].sheet_state = "hidden"
    wb["Lists"].sheet_state = "hidden"
    wb.active = wb.sheetnames.index("Dashboard")
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass
    return wb


def compute_headline_kpis(data):
    """Cross-check KPIs in pandas (Excel computes the live versions on open)."""
    inc = data["incidents"]
    act = data["activity"]
    today = REPORT_ANCHOR
    recordable = inc["Type"].isin(RECORDABLE_TYPES).sum()
    lti = (inc["Type"] == "Lost Time Injury").sum()
    manhours = act["ManHours"].sum()
    near = act["NearMisses"].sum()
    insp = (act["InspDone"] * act["InspScore"]).sum() / max(1, act["InspDone"].sum())
    train = act["TrainCompleted"].sum() / max(1, act["TrainAssigned"].sum())
    lti_dates = inc.loc[inc["Type"] == "Lost Time Injury", "Date"]
    days_since = (today - lti_dates.max()).days if len(lti_dates) else None
    actions = data["actions"]
    open_actions = (actions["Status"] != "Closed").sum()
    env = data["environmental"]
    oks = sum(1 for v in env["PM10"] if v <= 70) + sum(1 for v in env["pH"] if 6 <= v <= 9) \
        + sum(1 for v in env["WAD_CN"] if v <= 50)
    env_pct = oks / (3 * len(env))
    return {
        "Total incidents": int(len(inc)),
        "Recordables": int(recordable),
        "Lost Time Injuries": int(lti),
        "Near misses": int(near),
        "Man-hours": int(manhours),
        "TRIFR": round(recordable / manhours * 1_000_000, 2),
        "LTIFR": round(lti / manhours * 1_000_000, 2),
        "Inspection score": f"{insp*100:.1f}%",
        "Training completion": f"{train*100:.1f}%",
        "Environmental compliance": f"{env_pct*100:.1f}%",
        "Open corrective actions": int(open_actions),
        "Days since last LTI": days_since,
    }


def main():
    import os
    print("=" * 74)
    print(f"  {COMPANY_NAME}  |  HSE MANAGEMENT DASHBOARD GENERATOR")
    print("=" * 74)
    print(f"  Seeding {MONTHS_OF_HISTORY} months of sample data (numpy seed={RNG_SEED}) ...")
    data = generate_sample_data()
    print(f"    incidents={len(data['incidents'])}, activity rows={len(data['activity'])}, "
          f"actions={len(data['actions'])}, risks={len(data['risks'])}")

    print("  Building workbook ...")
    wb = build_workbook(data)

    out_path = os.path.abspath(OUTPUT_FILE)
    wb.save(out_path)

    print("\n  SHEETS CREATED:")
    for i, name in enumerate(wb.sheetnames, 1):
        state = "" if wb[name].sheet_state == "visible" else "  (hidden helper)"
        print(f"    {i:2d}. {name}{state}")

    print("\n  HEADLINE KPIs (pandas cross-check; Excel recomputes live on open):")
    for k, v in compute_headline_kpis(data).items():
        print(f"    - {k:28s}: {v}")

    print("\n  OUTPUT FILE:")
    print(f"    {out_path}")
    print("\n  Open in Excel. Edit only Incident_Register + Activity_Log; everything")
    print("  else (KPIs, traffic lights, charts) recalculates automatically.")
    print("=" * 74)


if __name__ == "__main__":
    main()
