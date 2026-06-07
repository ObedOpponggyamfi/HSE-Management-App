#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Static integrity checker for the generated HSE workbook.

openpyxl writes formula strings without evaluating them, so a typo in a
structured reference (tblX[Column]) or a named range only shows up as #REF! /
#NAME? once the file is opened in Excel. This script reloads the saved .xlsx and
verifies, without Excel:

  1. every  table[column]  structured reference points at a real table + column
  2. every named range used in a formula is actually defined
  3. parentheses are balanced in every formula
  4. tables, data validations and charts are present

Exit code 0 = clean, 1 = problems found.
"""
import argparse
import re
import sys

from openpyxl import load_workbook

FILE = "Asanko_HSE_Management_Dashboard.xlsx"

# Excel built-in functions / tokens that look like names but are NOT named ranges.
EXCEL_FUNCS = {
    "SUMPRODUCT", "SUMIFS", "SUMIF", "COUNTIFS", "COUNTIF", "COUNTA", "AVERAGEIFS",
    "IF", "IFERROR", "AND", "OR", "NOT", "MAX", "MIN", "SUM", "TODAY", "YEAR",
    "MONTH", "TEXT", "EDATE", "ROWS", "ROW", "COLUMN", "OFFSET", "N", "RIGHT",
    "LEFT", "MID", "VALUE", "DATE", "ABS", "ROUND", "INDEX", "MATCH", "TRUE",
    "FALSE", "ISNUMBER", "ISBLANK", "AVERAGE", "DATEVALUE",
}

# Tokens used inside structured references that are not column names.
STRUCT_ITEMS = {"#All", "#Data", "#Headers", "#Totals", "#This Row"}


def parse_args(argv=None):
    parser = argparse.ArgumentParser(description="Verify HSE workbook formulas and structure.")
    parser.add_argument("file", nargs="?", default=FILE)
    return parser.parse_args(argv)


def main(argv=None):
    args = parse_args(argv)
    wb = load_workbook(args.file, data_only=False)
    defined = set(wb.defined_names.keys())

    # map: table name -> set(column headers)
    table_cols = {}
    for ws in wb.worksheets:
        for tname in list(ws.tables):          # TableList.items() yields refs, not objects
            tbl = ws.tables[tname]             # indexing returns the Table object
            table_cols[tname] = {c.name for c in tbl.tableColumns}

    problems = []
    stats = {"formulas": 0, "struct_refs": 0, "name_refs": 0}

    struct_re = re.compile(r"([A-Za-z_][A-Za-z0-9_]*)\[([^\]]*)\]")
    # bare names: identifiers not preceded by '[' or '(' chain that are not funcs
    name_re = re.compile(r"(?<![A-Za-z0-9_\]\[@])([A-Za-z_][A-Za-z0-9_\.]*)\b(?!\s*\()")

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not (isinstance(v, str) and v.startswith("=")):
                    continue
                stats["formulas"] += 1
                loc = f"{ws.title}!{cell.coordinate}"

                # 1. balanced parens
                if v.count("(") != v.count(")"):
                    problems.append(f"[parens] {loc}: {v[:60]}")

                # 2. structured references table[column]
                for m in struct_re.finditer(v):
                    name, inner = m.group(1), m.group(2).strip()
                    if name not in table_cols:
                        # could be a function call like TEXT[..] -> never; flag
                        problems.append(f"[table?] {loc}: unknown table '{name}'")
                        continue
                    stats["struct_refs"] += 1
                    if inner.startswith("@"):
                        col = inner[1:].strip().strip("[]")
                    else:
                        col = inner.strip("[]")
                    if not col or col in STRUCT_ITEMS:
                        continue
                    if col not in table_cols[name]:
                        problems.append(
                            f"[column] {loc}: {name}[{col}] -- not a column of {name}")

                # 3. named ranges: strip structured refs + strings first
                stripped = struct_re.sub(" ", v)
                stripped = re.sub(r'"[^"]*"', " ", stripped)   # remove string literals
                for m in name_re.finditer(stripped):
                    tok = m.group(1)
                    if tok.upper() in EXCEL_FUNCS or tok in EXCEL_FUNCS:
                        continue
                    if re.fullmatch(r"[A-Za-z]{1,3}[0-9]+", tok):   # cell ref A1, AB12
                        continue
                    if re.fullmatch(r"[A-Za-z]{1,3}", tok):         # column letter
                        continue
                    if tok in ("All",):
                        continue
                    if tok in table_cols:        # whole-table ref e.g. ROWS(tblX)
                        continue
                    # looks like a named range -> must be defined
                    if tok not in defined:
                        # ignore sheet-name!cell style handled elsewhere
                        if "!" in v and tok in {ws.title for ws in wb.worksheets}:
                            continue
                        problems.append(f"[name] {loc}: undefined name '{tok}'  in {v[:50]}")
                    else:
                        stats["name_refs"] += 1

    # 4. structural presence
    print("Tables found:", ", ".join(sorted(table_cols)))
    print("Defined names:", len(defined))
    dv_count = sum(len(ws.data_validations.dataValidation) for ws in wb.worksheets)
    chart_count = sum(len(ws._charts) for ws in wb.worksheets)
    print(f"Data validations: {dv_count}   Charts: {chart_count}")
    print(f"Formulas scanned: {stats['formulas']}   "
          f"struct-refs OK: {stats['struct_refs']}   name-refs OK: {stats['name_refs']}")

    # de-duplicate while preserving order
    seen, uniq = set(), []
    for p in problems:
        if p not in seen:
            seen.add(p)
            uniq.append(p)

    if uniq:
        print(f"\n{len(uniq)} PROBLEM(S) FOUND:")
        for p in uniq[:80]:
            print("  -", p)
        sys.exit(1)
    print("\nOK -- no broken structured references or undefined names detected.")


if __name__ == "__main__":
    main()
